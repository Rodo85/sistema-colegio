from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.csrf import csrf_exempt
from django.utils.encoding import smart_str
from django.utils import timezone
from django.db.models import Count, Q, Value
from django.db.models.functions import Coalesce
from config_institucional.models import Nivel
from catalogos.models import CursoLectivo, Seccion, Subgrupo, Especialidad
from core.models import Institucion
from .models import Estudiante, EstudianteInstitucion, MatriculaAcademica, PlantillaImpresionMatricula
from dal import autocomplete
import json
import io
import qrcode
from io import BytesIO
import base64

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

@login_required
@permission_required('matricula.access_consulta_estudiante', raise_exception=True)
def consulta_estudiante(request):
    estudiante = None
    matricula = None
    encargados = []
    curso_lectivo = None
    identificacion = ''
    error = ''
    institucion = None
    edad_estudiante = ""
    
    # Obtener todos los cursos lectivos disponibles para el select
    cursos_lectivos = CursoLectivo.objects.all().order_by('-anio')
    
    # Obtener instituciones disponibles solo para superusuario
    instituciones = []
    if request.user.is_superuser:
        instituciones = Institucion.objects.all().order_by('nombre')
    
    # Obtener la plantilla de impresión específica de la institución
    plantilla = None
    if request.user.is_superuser:
        # Superusuario: obtener plantilla según institución seleccionada
        institucion_id = request.POST.get('institucion') if request.method == 'POST' else None
        if institucion_id:
            try:
                institucion_temp = Institucion.objects.get(pk=institucion_id)
                plantilla = PlantillaImpresionMatricula.objects.filter(institucion=institucion_temp).first()
            except:
                pass
    else:
        # Usuario normal: obtener plantilla de su institución activa
        institucion_id = getattr(request, 'institucion_activa_id', None)
        if institucion_id:
            try:
                plantilla = PlantillaImpresionMatricula.objects.filter(institucion_id=institucion_id).first()
            except:
                pass
    
    if request.method == 'POST':
        accion = request.POST.get('accion', 'buscar')
        curso_lectivo_id = request.POST.get('curso_lectivo')
        identificacion = request.POST.get('identificacion', '').strip()
        institucion_id = request.POST.get('institucion')

        if accion == 'limpiar':
            if curso_lectivo_id:
                curso_lectivo = CursoLectivo.objects.filter(pk=curso_lectivo_id).first()
            if request.user.is_superuser and institucion_id:
                institucion = Institucion.objects.filter(pk=institucion_id).first()
            elif not request.user.is_superuser:
                institucion_activa_id = getattr(request, 'institucion_activa_id', None)
                if institucion_activa_id:
                    institucion = Institucion.objects.filter(pk=institucion_activa_id).first()

            identificacion = ''
            estudiante = None
            matricula = None
            encargados = []
            edad_estudiante = ""
            error = ''
        else:
            if not curso_lectivo_id or not identificacion:
                error = 'Debe seleccionar un curso lectivo e ingresar la identificación del estudiante.'
            elif request.user.is_superuser and not institucion_id:
                error = 'Debe seleccionar una institución.'
            else:
                try:
                    curso_lectivo = CursoLectivo.objects.get(pk=curso_lectivo_id)
                    
                    # Determinar la institución según el tipo de usuario
                    if request.user.is_superuser:
                        institucion = Institucion.objects.get(pk=institucion_id)
                    else:
                        # Usuario normal: usar institución activa
                        institucion_id = getattr(request, 'institucion_activa_id', None)
                        if not institucion_id:
                            error = 'No se pudo determinar la institución activa.'
                            return render(request, 'matricula/consulta_estudiante.html', {
                                'error': error,
                                'cursos_lectivos': cursos_lectivos,
                                'instituciones': instituciones,
                                'es_superusuario': request.user.is_superuser,
                                'plantilla': plantilla,
                            })
                        institucion = Institucion.objects.get(pk=institucion_id)
                    
                    # Buscar estudiante por identificación
                    try:
                        estudiante = Estudiante.objects.get(identificacion=identificacion)
                        
                        # Verificar que tenga relación activa con la institución
                        relacion_activa = estudiante.instituciones_estudiante.filter(
                            institucion=institucion,
                            estado='activo'
                        ).exists()
                        
                        if not relacion_activa:
                            error = f'El estudiante {identificacion} no pertenece a la institución seleccionada o no está activo en ella.'
                            estudiante = None
                    except Estudiante.DoesNotExist:
                        error = f'No se encontró ningún estudiante con la identificación {identificacion}.'
                        estudiante = None
                    
                    # Buscar matrícula activa para el curso seleccionado solo si estudiante es válido
                    matricula = None
                    if estudiante:
                        matricula = MatriculaAcademica.objects.filter(
                            estudiante=estudiante,
                            curso_lectivo=curso_lectivo,
                            estado__iexact='activo'
                        ).first()
                    
                    if matricula and estudiante:
                        # Si hay matrícula activa, obtener encargados
                        encargados = estudiante.encargadoestudiante_set.select_related(
                            'persona_contacto', 'parentesco'
                        ).all()
                        
                        # Calcular edad del estudiante
                        if estudiante.fecha_nacimiento:
                            from datetime import date
                            today = date.today()
                            years = today.year - estudiante.fecha_nacimiento.year
                            months = today.month - estudiante.fecha_nacimiento.month
                            days = today.day - estudiante.fecha_nacimiento.day

                            # Ajustar meses y años si todavía no cumple años este mes
                            if days < 0:
                                months -= 1

                            if months < 0:
                                years -= 1
                                months += 12

                            # Formatear la edad
                            if years == 0:
                                edad_estudiante = f"{months} meses"
                            elif months == 0:
                                edad_estudiante = f"{years} años"
                            else:
                                edad_estudiante = f"{years} años y {months} meses"
                    else:
                        # No hay matrícula activa, mostrar error
                        estudiante = None
                        institucion = None
                        encargados = []
                        error = f'No existe matrícula activa para el estudiante con identificación {identificacion} en el curso lectivo {curso_lectivo.nombre}.'
                        
                except CursoLectivo.DoesNotExist:
                    error = 'El curso lectivo seleccionado no existe.'
                except Institucion.DoesNotExist:
                    error = 'La institución seleccionada no existe.'
                except Estudiante.DoesNotExist:
                    error = f'No se encontró ningún estudiante con la identificación {identificacion} en la institución {institucion.nombre}.'
                except Exception as e:
                    # Log del error para debugging
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Error inesperado en consulta_estudiante: {e}")
                    error = 'Error interno del sistema. Contacte al administrador.'
    
    context = {
        'estudiante': estudiante,
        'matricula': matricula,
        'encargados': encargados,
        'curso_lectivo': curso_lectivo,
        'identificacion': identificacion,
        'error': error,
        'institucion': institucion,
        'edad_estudiante': edad_estudiante,
        'cursos_lectivos': cursos_lectivos,
        'instituciones': instituciones,
        'es_superusuario': request.user.is_superuser,
        'plantilla': plantilla,
    }
    
    return render(request, 'matricula/consulta_estudiante.html', context)


@login_required
@permission_required('matricula.access_reporte_matricula', raise_exception=True)
def reporte_matricula(request):
    cursos_lectivos = CursoLectivo.objects.all().order_by('-anio')
    curso_lectivo_id = request.GET.get('curso_lectivo')
    curso_lectivo = None

    if not curso_lectivo_id and cursos_lectivos.exists():
        curso_lectivo = cursos_lectivos.first()
        curso_lectivo_id = str(curso_lectivo.pk)
    elif curso_lectivo_id:
        try:
            curso_lectivo = cursos_lectivos.get(pk=curso_lectivo_id)
        except CursoLectivo.DoesNotExist:
            curso_lectivo = None
            curso_lectivo_id = None

    instituciones = []
    institucion = None
    institucion_id = None
    error = ''

    if request.user.is_superuser:
        instituciones = Institucion.objects.all().order_by('nombre')
        institucion_id = request.GET.get('institucion')
        if institucion_id:
            try:
                institucion = instituciones.get(pk=institucion_id)
            except Institucion.DoesNotExist:
                institucion = None
                institucion_id = None
                error = 'La institución seleccionada no existe.'
    else:
        institucion_id = getattr(request, 'institucion_activa_id', None)
        if institucion_id:
            try:
                institucion = Institucion.objects.get(pk=institucion_id)
            except Institucion.DoesNotExist:
                institucion = None
                error = 'No se pudo determinar la institución activa.'
        else:
            error = 'No se pudo determinar la institución activa.'

    resumen = {
        'total': 0,
        'hombres': 0,
        'mujeres': 0,
        'pn': 0,
        'pr': 0,
        'otros_generos': 0,
    }
    genero_tipo = []
    niveles = []
    especialidades = []
    estados = []
    sin_matricula = {
        'total': 0,
        'pn': 0,
        'pr': 0,
    }
    sin_matricula_genero = []

    if curso_lectivo_id and institucion_id and not error:
        base_qs = (
            MatriculaAcademica.objects.filter(
                curso_lectivo_id=curso_lectivo_id,
                institucion_id=institucion_id,
                estado__iexact='activo'
            )
            .select_related(
                'estudiante__sexo',
                'nivel',
                'especialidad__especialidad'
            )
        )

        aggregated = base_qs.aggregate(
            total=Count('id', distinct=True),
            hombres=Count('id', filter=Q(estudiante__sexo__codigo='M'), distinct=True),
            mujeres=Count('id', filter=Q(estudiante__sexo__codigo='F'), distinct=True),
            pn=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PN), distinct=True),
            pr=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PR), distinct=True),
        )

        for key, value in aggregated.items():
            resumen[key] = value or 0

        resumen['otros_generos'] = (
            resumen['total'] - resumen['hombres'] - resumen['mujeres']
        )

        genero_tipo = list(
            base_qs.values(
                'estudiante__sexo__codigo',
                'estudiante__sexo__nombre'
            ).annotate(
                total=Count('id', distinct=True),
                pn=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PN), distinct=True),
                pr=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PR), distinct=True),
            ).order_by('estudiante__sexo__nombre')
        )

        niveles = list(
            base_qs.values(
                'nivel__id',
                'nivel__nombre',
                'nivel__numero'
            ).annotate(
                total=Count('id', distinct=True),
                pn=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PN), distinct=True),
                pr=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PR), distinct=True),
            ).order_by('nivel__numero')
        )

        especialidades = list(
            base_qs.annotate(
                especialidad_nombre=Coalesce(
                    'especialidad__especialidad__nombre',
                    Value('SIN ESPECIALIDAD')
                )
            )
            .filter(
                Q(especialidad__isnull=True) | Q(especialidad__activa=True)
            )
            .values('especialidad_nombre')
            .annotate(
                total=Count('id', distinct=True),
                pn=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PN), distinct=True),
                pr=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PR), distinct=True),
            )
            .order_by('especialidad_nombre')
        )

        estados = list(
            MatriculaAcademica.objects.filter(
                curso_lectivo_id=curso_lectivo_id,
                institucion_id=institucion_id,
            )
            .values('estado')
            .annotate(
                total=Count('id', distinct=True),
                pn=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PN), distinct=True),
                pr=Count('id', filter=Q(estudiante__tipo_estudiante=Estudiante.PR), distinct=True),
            )
            .order_by('estado')
        )

        estudiantes_activos = Estudiante.objects.filter(
            instituciones_estudiante__institucion_id=institucion_id,
            instituciones_estudiante__estado=EstudianteInstitucion.ACTIVO
        ).distinct()

        activos_sin_matricula = estudiantes_activos.exclude(
            matriculas_academicas__curso_lectivo_id=curso_lectivo_id,
            matriculas_academicas__estado__iexact='activo'
        )

        sin_matricula['total'] = activos_sin_matricula.count()
        sin_por_tipo = activos_sin_matricula.values('tipo_estudiante').annotate(
            total=Count('id', distinct=True)
        )
        for item in sin_por_tipo:
            codigo = item['tipo_estudiante']
            if codigo == Estudiante.PN:
                sin_matricula['pn'] = item['total']
            elif codigo == Estudiante.PR:
                sin_matricula['pr'] = item['total']

        sin_matricula_genero = list(
            activos_sin_matricula.values(
                'sexo__codigo',
                'sexo__nombre'
            ).annotate(
                total=Count('id', distinct=True),
            ).order_by('sexo__nombre')
        )

    tipo_estudiante_labels = dict(Estudiante.TIPO_CHOICES)

    context = {
        'cursos_lectivos': cursos_lectivos,
        'curso_lectivo': curso_lectivo,
        'curso_lectivo_id': curso_lectivo_id,
        'instituciones': instituciones,
        'institucion': institucion,
        'institucion_id': institucion_id,
        'es_superusuario': request.user.is_superuser,
        'error': error,
        'resumen': resumen,
        'genero_tipo': genero_tipo,
        'niveles': niveles,
        'especialidades': especialidades,
        'estados': estados,
        'sin_matricula': sin_matricula,
        'sin_matricula_genero': sin_matricula_genero,
        'tipo_estudiante_labels': tipo_estudiante_labels,
    }

    return render(request, 'matricula/reporte_matricula.html', context)


@login_required
def comprobante_matricula(request):
    """
    Renderiza un comprobante de matrícula imprimible para un estudiante y curso lectivo dados.
    Requiere: curso_lectivo_id, identificacion y (si es superusuario) institucion_id via querystring.
    """
    error = ''
    estudiante = None
    matricula = None
    institucion = None
    plantilla = None
    edad_estudiante = ''
    contacto_principal = None

    curso_lectivo_id = request.GET.get('curso_lectivo_id')
    identificacion = (request.GET.get('identificacion') or '').strip()
    institucion_id = request.GET.get('institucion_id') if request.user.is_superuser else getattr(request, 'institucion_activa_id', None)

    if not curso_lectivo_id or not identificacion:
        return HttpResponse('Parámetros insuficientes', status=400)

    try:
        curso_lectivo = CursoLectivo.objects.get(pk=curso_lectivo_id)

        # Determinar institución
        if request.user.is_superuser:
            if not institucion_id:
                return HttpResponse('Institución requerida para superusuarios', status=400)
            institucion = Institucion.objects.get(pk=institucion_id)
        else:
            if not institucion_id:
                return HttpResponse('No se pudo determinar la institución activa', status=400)
            institucion = Institucion.objects.get(pk=institucion_id)

        # Plantilla de impresión para el encabezado
        try:
            plantilla = PlantillaImpresionMatricula.objects.filter(institucion=institucion).first()
        except Exception:
            plantilla = None

        # Estudiante y matrícula activa
        # El estudiante usa tabla intermedia EstudianteInstitucion
        estudiante = Estudiante.objects.filter(
            identificacion=identificacion,
            instituciones_estudiante__institucion=institucion,
            instituciones_estudiante__estado='activo'
        ).first()
        
        if not estudiante:
            return HttpResponse('Estudiante no encontrado en la institución indicada', status=404)
        
        matricula = MatriculaAcademica.objects.filter(
            estudiante=estudiante,
            curso_lectivo=curso_lectivo,
            estado__iexact='activo'
        ).select_related('nivel', 'especialidad__especialidad').first()

        if not matricula:
            return HttpResponse('No existe matrícula activa para este estudiante y curso lectivo', status=404)

        # Edad en años y meses
        if estudiante.fecha_nacimiento:
            from datetime import date
            today = date.today()
            years = today.year - estudiante.fecha_nacimiento.year
            months = today.month - estudiante.fecha_nacimiento.month
            if today.day < estudiante.fecha_nacimiento.day:
                months -= 1
            if months < 0:
                years -= 1
                months += 12
            if years == 0:
                edad_estudiante = f"{months} meses"
            elif months == 0:
                edad_estudiante = f"{years} años"
            else:
                edad_estudiante = f"{years} años y {months} meses"

        # Encargado principal
        contacto_principal = (
            estudiante.encargadoestudiante_set
            .select_related('persona_contacto', 'parentesco')
            .filter(principal=True)
            .first()
        )
        if not contacto_principal:
            contacto_principal = (
                estudiante.encargadoestudiante_set
                .select_related('persona_contacto', 'parentesco')
                .first()
            )

        context = {
            'estudiante': estudiante,
            'matricula': matricula,
            'curso_lectivo': curso_lectivo,
            'institucion': institucion,
            'plantilla': plantilla,
            'edad_estudiante': edad_estudiante,
            'contacto_principal': contacto_principal,
        }
        return render(request, 'matricula/comprobante_matricula.html', context)
    except CursoLectivo.DoesNotExist:
        return HttpResponse('Curso lectivo no encontrado', status=404)
    except Institucion.DoesNotExist:
        return HttpResponse('Institución no encontrada', status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en comprobante_matricula: {str(e)}")
        return HttpResponse('Error interno del sistema', status=500)


@login_required
@permission_required('matricula.print_pas_estudiante', raise_exception=True)
def pas_estudiante(request):
    """
    Genera un PAS (Pase de Asistencia del Estudiante) en formato PDF compacto (11.5cm x 19cm).
    Incluye: foto, código QR, escudo institucional, datos básicos del estudiante y encargado principal.
    """
    curso_lectivo_id = request.GET.get('curso_lectivo_id')
    identificacion = (request.GET.get('identificacion') or '').strip()
    institucion_id = request.GET.get('institucion_id') if request.user.is_superuser else getattr(request, 'institucion_activa_id', None)

    if not curso_lectivo_id or not identificacion:
        return HttpResponse('Parámetros insuficientes', status=400)

    try:
        curso_lectivo = CursoLectivo.objects.get(pk=curso_lectivo_id)

        # Determinar institución
        if request.user.is_superuser:
            if not institucion_id:
                return HttpResponse('Institución requerida para superusuarios', status=400)
            institucion = Institucion.objects.get(pk=institucion_id)
        else:
            if not institucion_id:
                return HttpResponse('No se pudo determinar la institución activa', status=400)
            institucion = Institucion.objects.get(pk=institucion_id)

        # Buscar estudiante
        estudiante = Estudiante.objects.filter(
            identificacion=identificacion,
            instituciones_estudiante__institucion=institucion,
            instituciones_estudiante__estado='activo'
        ).first()
        
        if not estudiante:
            return HttpResponse('Estudiante no encontrado en la institución indicada', status=404)
        
        matricula = MatriculaAcademica.objects.filter(
            estudiante=estudiante,
            curso_lectivo=curso_lectivo,
            estado__iexact='activo'
        ).select_related('nivel', 'especialidad__especialidad').first()

        if not matricula:
            return HttpResponse('No existe matrícula activa para este estudiante y curso lectivo', status=404)

        # Generar código QR con la identificación del estudiante
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=1,
        )
        qr.add_data(estudiante.identificacion)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Convertir QR a base64
        buffered = BytesIO()
        qr_img.save(buffered, format="PNG")
        qr_base64 = base64.b64encode(buffered.getvalue()).decode()

        # Encargado principal
        contacto_principal = (
            estudiante.encargadoestudiante_set
            .select_related('persona_contacto', 'parentesco')
            .filter(principal=True)
            .first()
        )
        if not contacto_principal:
            contacto_principal = (
                estudiante.encargadoestudiante_set
                .select_related('persona_contacto', 'parentesco')
                .first()
            )

        context = {
            'estudiante': estudiante,
            'matricula': matricula,
            'institucion': institucion,
            'qr_code_base64': qr_base64,
            'contacto_principal': contacto_principal,
        }
        return render(request, 'matricula/pas_estudiante.html', context)
    except CursoLectivo.DoesNotExist:
        return HttpResponse('Curso lectivo no encontrado', status=404)
    except Institucion.DoesNotExist:
        return HttpResponse('Institución no encontrada', status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en pas_estudiante: {str(e)}")
        return HttpResponse('Error interno del sistema', status=500)


@login_required
@permission_required('matricula.print_pas_estudiante', raise_exception=True)
def pas_seccion(request):
    """
    Genera un PDF con múltiples PAS para todos los estudiantes de una sección o subgrupo.
    Parámetros: curso_lectivo_id, seccion_id O subgrupo_id, institucion_id (si superuser)
    """
    curso_lectivo_id = request.GET.get('curso_lectivo_id')
    seccion_id = request.GET.get('seccion_id')
    subgrupo_id = request.GET.get('subgrupo_id')
    institucion_id = request.GET.get('institucion_id') if request.user.is_superuser else getattr(request, 'institucion_activa_id', None)

    if not curso_lectivo_id or (not seccion_id and not subgrupo_id):
        return HttpResponse('Parámetros insuficientes: requiere curso_lectivo_id y (seccion_id o subgrupo_id)', status=400)

    try:
        curso_lectivo = CursoLectivo.objects.get(pk=curso_lectivo_id)

        # Determinar institución
        if request.user.is_superuser:
            if not institucion_id:
                return HttpResponse('Institución requerida para superusuarios', status=400)
            institucion = Institucion.objects.get(pk=institucion_id)
        else:
            if not institucion_id:
                return HttpResponse('No se pudo determinar la institución activa', status=400)
            institucion = Institucion.objects.get(pk=institucion_id)

        # Filtrar matrículas según sección o subgrupo
        filtros = {
            'curso_lectivo': curso_lectivo,
            'institucion': institucion,
            'estado__iexact': 'activo'
        }

        if subgrupo_id:
            filtros['subgrupo_id'] = subgrupo_id
            grupo_nombre = f"Subgrupo {Subgrupo.objects.get(pk=subgrupo_id)}"
        elif seccion_id:
            filtros['seccion_id'] = seccion_id
            grupo_nombre = f"Sección {Seccion.objects.get(pk=seccion_id)}"

        # Obtener estudiantes ordenados alfabéticamente
        matriculas = MatriculaAcademica.objects.filter(**filtros).select_related(
            'estudiante', 'nivel', 'especialidad__especialidad'
        ).order_by(
            'estudiante__primer_apellido',
            'estudiante__segundo_apellido',
            'estudiante__nombres'
        )

        if not matriculas.exists():
            return HttpResponse(f'No se encontraron estudiantes en {grupo_nombre}', status=404)

        # Preparar datos para cada estudiante
        estudiantes_data = []
        for matricula in matriculas:
            # Generar código QR
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=1,
            )
            qr.add_data(matricula.estudiante.identificacion)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            buffered = BytesIO()
            qr_img.save(buffered, format="PNG")
            qr_base64 = base64.b64encode(buffered.getvalue()).decode()

            # Encargado principal
            contacto_principal = (
                matricula.estudiante.encargadoestudiante_set
                .select_related('persona_contacto', 'parentesco')
                .filter(principal=True)
                .first()
            )
            if not contacto_principal:
                contacto_principal = (
                    matricula.estudiante.encargadoestudiante_set
                    .select_related('persona_contacto', 'parentesco')
                    .first()
                )

            estudiantes_data.append({
                'estudiante': matricula.estudiante,
                'matricula': matricula,
                'qr_code_base64': qr_base64,
                'contacto_principal': contacto_principal,
            })

        context = {
            'institucion': institucion,
            'estudiantes_data': estudiantes_data,
            'grupo_nombre': grupo_nombre,
            'total_estudiantes': len(estudiantes_data),
        }
        return render(request, 'matricula/pas_seccion.html', context)
    except (CursoLectivo.DoesNotExist, Institucion.DoesNotExist, Seccion.DoesNotExist, Subgrupo.DoesNotExist):
        return HttpResponse('Recurso no encontrado', status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en pas_seccion: {str(e)}")
        return HttpResponse('Error interno del sistema', status=500)

@login_required
@permission_required('matricula.access_reporte_pas_seccion', raise_exception=True)
def reporte_pas_seccion(request):
    """
    Interfaz para seleccionar y generar PAS por sección o subgrupo.
    """
    cursos_lectivos = CursoLectivo.objects.all().order_by('-anio')
    niveles = Nivel.objects.all().order_by('numero')
    
    # Obtener instituciones si es superusuario
    instituciones = []
    institucion = None
    if request.user.is_superuser:
        instituciones = Institucion.objects.all().order_by('nombre')
    else:
        institucion_id = getattr(request, 'institucion_activa_id', None)
        if institucion_id:
            institucion = Institucion.objects.get(pk=institucion_id)
    
    context = {
        'cursos_lectivos': cursos_lectivos,
        'niveles': niveles,
        'instituciones': instituciones,
        'institucion': institucion,
        'es_superusuario': request.user.is_superuser,
    }
    
    return render(request, 'matricula/reporte_pas_seccion.html', context)


@csrf_exempt
@login_required
def get_especialidades_disponibles(request):
    """
    Vista AJAX para obtener las especialidades disponibles según el curso lectivo
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        curso_lectivo_id = request.POST.get('curso_lectivo_id')
        if not curso_lectivo_id:
            return JsonResponse({'success': False, 'error': 'ID de curso lectivo requerido'})
        
        # Obtener la institución del usuario
        institucion_id = getattr(request, 'institucion_activa_id', None)
        if not institucion_id:
            return JsonResponse({'success': False, 'error': 'No se pudo determinar la institución'})
        
        # Obtener curso lectivo (es GLOBAL, no tiene institucion_id)
        from core.models import Institucion
        curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
        institucion = Institucion.objects.get(id=institucion_id)
        
        # Obtener especialidades disponibles
        especialidades_disponibles = MatriculaAcademica.get_especialidades_disponibles(
            institucion=institucion,
            curso_lectivo=curso_lectivo
        )
        
        # Preparar datos para JSON
        especialidades_data = []
        for ecl in especialidades_disponibles:
            # ecl es EspecialidadCursoLectivo; extraer la especialidad real
            if hasattr(ecl, 'especialidad') and ecl.especialidad:
                especialidades_data.append({
                    'id': ecl.especialidad.id,
                    'nombre': ecl.especialidad.nombre,
                    'modalidad': getattr(getattr(ecl.especialidad, 'modalidad', None), 'nombre', '')
                })
        
        # Debug: verificar configuraciones existentes
        from config_institucional.models import EspecialidadCursoLectivo
        configuraciones_count = EspecialidadCursoLectivo.objects.filter(
            institucion=institucion,
            curso_lectivo=curso_lectivo,
            activa=True
        ).count()
        
        return JsonResponse({
            'success': True,
            'especialidades': especialidades_data,
            'curso_lectivo': curso_lectivo.nombre,
            'debug': {
                'institucion_id': institucion.id,
                'institucion_nombre': institucion.nombre,
                'curso_lectivo_id': curso_lectivo.id,
                'configuraciones_activas': configuraciones_count,
                'total_especialidades': len(especialidades_data)
            }
        })
        
    except CursoLectivo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Curso lectivo no encontrado'})
    except Institucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Institución no encontrada'})
    except Exception as e:
        # Log del error para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error inesperado en get_especialidades_disponibles: {e}")
        return JsonResponse({'success': False, 'error': 'Error interno del sistema'})



# ════════════════════════════════════════════════════════════════
#                    DJANGO AUTOCOMPLETE LIGHT (DAL)
# ════════════════════════════════════════════════════════════════


class EspecialidadAutocomplete(autocomplete.Select2QuerySetView):
    """
    Autocompletado para Especialidad que filtra por Curso Lectivo, Nivel e institución.
    Solo muestra especialidades para niveles 10, 11, 12.
    Busca directamente en EspecialidadCursoLectivo.
    Forward: curso_lectivo, nivel → especialidad
    """
    def get_queryset(self):
        # DEBUG: Imprimir información de debug
        print("🔥 EspecialidadAutocomplete.get_queryset() llamado")
        
        if not self.request.user.is_authenticated:
            print("❌ Usuario no autenticado")
            return Especialidad.objects.none()
        
        print(f"👤 Usuario: {self.request.user.email} (superuser: {self.request.user.is_superuser})")
        
        # Obtener institución del usuario
        institucion_id = getattr(self.request, 'institucion_activa_id', None)
        print(f"🏢 Institución activa ID: {institucion_id}")
        
        if not institucion_id:
            # Para superusuario, usar institución 1 por defecto
            if self.request.user.is_superuser:
                institucion_id = 1
                print(f"👑 Superusuario - usando institución por defecto: {institucion_id}")
            else:
                print("❌ No hay institución activa")
                return Especialidad.objects.none()
        
        # FILTRO POR CURSO LECTIVO Y NIVEL (forward) - BUSCAR EN EspecialidadCursoLectivo
        curso_lectivo_id = self.forwarded.get('curso_lectivo', None)
        nivel_id = self.forwarded.get('nivel', None)
        print(f"📅 Curso lectivo ID: {curso_lectivo_id}")
        print(f"📊 Nivel ID: {nivel_id}")
        print(f"📅 Forwarded completo: {self.forwarded}")
        
        # Inicializar qs como None
        qs = None
        
        # REQUIERE tanto curso lectivo como nivel
        if curso_lectivo_id and nivel_id:
            try:
                from config_institucional.models import EspecialidadCursoLectivo
                from catalogos.models import Nivel
                
                # Verificar que el curso lectivo existe (ahora es global)
                curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
                print(f"✅ Curso lectivo encontrado: {curso_lectivo.nombre}")
                
                # Verificar que el nivel existe
                nivel = Nivel.objects.get(id=nivel_id)
                print(f"✅ Nivel encontrado: {nivel}")
                
                # SOLO mostrar especialidades para niveles 10, 11, 12
                if nivel.numero not in [10, 11, 12]:
                    print(f"❌ Nivel {nivel.numero} ({nivel.nombre}) no requiere especialidad")
                    return Especialidad.objects.none()
                
                print(f"✅ Nivel {nivel.numero} ({nivel.nombre}) requiere especialidad")
                
                # Obtener especialidades configuradas y activas para este curso lectivo
                qs = EspecialidadCursoLectivo.objects.filter(
                    institucion_id=institucion_id,
                    curso_lectivo=curso_lectivo,
                    activa=True
                ).select_related('especialidad', 'especialidad__modalidad')
                
                # FILTRO ADICIONAL: Solo especialidades vinculadas a subgrupos de este nivel
                from config_institucional.models import SubgrupoCursoLectivo
                
                especialidades_ids = SubgrupoCursoLectivo.objects.filter(
                    curso_lectivo=curso_lectivo,
                    institucion_id=institucion_id,
                    subgrupo__seccion__nivel=nivel,
                    activa=True,
                    especialidad_curso__isnull=False
                ).values_list('especialidad_curso_id', flat=True).distinct()
                
                print(f"🔍 IDs de especialidades vinculadas al nivel {nivel.numero}: {list(especialidades_ids)}")
                
                # Filtrar para incluir solo las especialidades del nivel
                qs = qs.filter(id__in=especialidades_ids)
                
                print(f"🎯 Especialidades configuradas encontradas: {[ecl.especialidad.nombre if ecl.especialidad else 'N/A' for ecl in qs]}")
                print(f"📋 Especialidades filtradas por nivel: {[ecl.especialidad.nombre if ecl.especialidad else 'N/A' for ecl in qs]}")
                
            except (CursoLectivo.DoesNotExist, Nivel.DoesNotExist, ValueError) as e:
                print(f"❌ Error: {e}")
                return Especialidad.objects.none()
        else:
            # Sin curso lectivo o nivel, no mostrar especialidades
            if not curso_lectivo_id:
                print("❌ No hay curso lectivo seleccionado")
            if not nivel_id:
                print("❌ No hay nivel seleccionado")
            return Especialidad.objects.none()
        
        # Verificar que qs esté definido
        if qs is None:
            print("❌ qs no está definido")
            return Especialidad.objects.none()
        
        # Filtro por búsqueda
        if self.q:
            qs = qs.filter(especialidad__nombre__icontains=self.q)
            print(f"🔍 Filtrado por búsqueda '{self.q}': {[ecl.especialidad.nombre if ecl.especialidad else 'N/A' for ecl in qs]}")
        
        final_qs = qs.order_by('especialidad__nombre')
        print(f"🎯 RESULTADO FINAL: {[str(ecl) for ecl in final_qs]}")
        return final_qs


class SeccionAutocomplete(autocomplete.Select2QuerySetView):
    """
    Autocompletado para Sección que filtra por Curso Lectivo, Nivel, Especialidad e institución.
    Busca directamente en SeccionCursoLectivo.
    Forward: curso_lectivo, nivel, especialidad → seccion
    """
    def get_queryset(self):
        print("🔥 SeccionAutocomplete.get_queryset() llamado")
        
        if not self.request.user.is_authenticated:
            print("❌ Usuario no autenticado")
            return Seccion.objects.none()
        
        print(f"👤 Usuario: {self.request.user.email} (superuser: {self.request.user.is_superuser})")
        
        # Obtener institución del usuario
        institucion_id = getattr(self.request, 'institucion_activa_id', None)
        print(f"🏢 Institución activa ID: {institucion_id}")
        
        if not institucion_id:
            # Para superusuario, usar institución 1 por defecto
            if self.request.user.is_superuser:
                institucion_id = 1
                print(f"👑 Superusuario - usando institución por defecto: {institucion_id}")
            else:
                print("❌ No hay institución activa")
                return Seccion.objects.none()
        
        # FILTRO POR CURSO LECTIVO Y NIVEL (forward) - BUSCAR EN SeccionCursoLectivo
        curso_lectivo_id = self.forwarded.get('curso_lectivo', None)
        nivel_id = self.forwarded.get('nivel', None)
        especialidad_id = self.forwarded.get('especialidad', None)
        print(f"📅 Curso lectivo ID: {curso_lectivo_id}")
        print(f"📅 Nivel ID: {nivel_id}")
        print(f"🎓 Especialidad ID: {especialidad_id}")
        print(f"📅 Forwarded completo: {self.forwarded}")
        
        if curso_lectivo_id and nivel_id:
            try:
                from config_institucional.models import SeccionCursoLectivo, SubgrupoCursoLectivo
                from catalogos.models import Nivel
                
                # Verificar que el curso lectivo existe (ahora es global)
                curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
                print(f"✅ Curso lectivo encontrado: {curso_lectivo.nombre}")
                
                # Verificar que el nivel existe
                nivel = Nivel.objects.get(id=nivel_id)
                print(f"✅ Nivel encontrado: {nivel}")
                
                # SI HAY ESPECIALIDAD (niveles 10, 11, 12): filtrar por especialidad
                if especialidad_id:
                    print(f"🎓 Filtrando por ESPECIALIDAD: {especialidad_id}")
                    # Obtener secciones que tienen subgrupos con esta especialidad
                    secciones_con_especialidad = SubgrupoCursoLectivo.objects.filter(
                        institucion_id=institucion_id,
                        curso_lectivo=curso_lectivo,
                        activa=True,
                        especialidad_curso_id=especialidad_id
                    ).values_list('subgrupo__seccion_id', flat=True).distinct()
                    
                    print(f"🎯 Secciones con especialidad IDs: {list(secciones_con_especialidad)}")
                    
                    # Filtrar secciones por nivel y que tengan la especialidad
                    qs = Seccion.objects.filter(
                        id__in=secciones_con_especialidad,
                        nivel=nivel
                    )
                    print(f"📋 Secciones con especialidad para nivel {nivel.numero}: {[sec.numero for sec in qs]}")
                else:
                    # Sin especialidad: mostrar todas las secciones configuradas
                    print(f"📋 Sin especialidad - mostrando todas las secciones")
                    # Obtener secciones configuradas y activas para este curso lectivo
                    secciones_configuradas = SeccionCursoLectivo.objects.filter(
                        institucion_id=institucion_id,
                        curso_lectivo=curso_lectivo,
                        activa=True
                    ).values_list('seccion_id', flat=True)
                    
                    print(f"🎯 Secciones configuradas IDs: {list(secciones_configuradas)}")
                    
                    # Filtrar secciones por nivel
                    qs = Seccion.objects.filter(
                        id__in=secciones_configuradas,
                        nivel=nivel
                    )
                    print(f"📋 Secciones encontradas para nivel {nivel.numero}: {[sec.numero for sec in qs]}")
                
            except (CursoLectivo.DoesNotExist, Nivel.DoesNotExist, ValueError) as e:
                print(f"❌ Error: {e}")
                return Seccion.objects.none()
        else:
            # Sin curso lectivo, no mostrar secciones
            print("❌ No hay curso lectivo o nivel seleccionado")
            return Seccion.objects.none()
        
        # Filtro por búsqueda
        if self.q:
            qs = qs.filter(numero__icontains=self.q)
            print(f"🔍 Filtrado por búsqueda '{self.q}': {[f'Sección {s.numero}' for s in qs]}")
        
        final_qs = qs.order_by('nivel__numero', 'numero')
        print(f"🎯 RESULTADO FINAL: {[f'Sección {s.numero} - {s.nivel.nombre}' for s in final_qs]}")
        return final_qs


class SubgrupoAutocomplete(autocomplete.Select2QuerySetView):
    """
    Autocompletado para Subgrupo que filtra por Curso Lectivo, Sección, Especialidad e institución.
    Busca directamente en SubgrupoCursoLectivo.
    Forward: curso_lectivo, seccion, especialidad → subgrupo
    """
    def get_queryset(self):
        print("🔥 SubgrupoAutocomplete.get_queryset() llamado")
        
        if not self.request.user.is_authenticated:
            print("❌ Usuario no autenticado")
            return Subgrupo.objects.none()
        
        print(f"👤 Usuario: {self.request.user.email} (superuser: {self.request.user.is_superuser})")
        
        # Obtener institución del usuario
        institucion_id = getattr(self.request, 'institucion_activa_id', None)
        print(f"🏢 Institución activa ID: {institucion_id}")
        
        if not institucion_id:
            # Para superusuario, usar institución 1 por defecto
            if self.request.user.is_superuser:
                institucion_id = 1
                print(f"👑 Superusuario - usando institución por defecto: {institucion_id}")
            else:
                print("❌ No hay institución activa")
                return Subgrupo.objects.none()
        
        # FILTRO POR CURSO LECTIVO, SECCIÓN Y ESPECIALIDAD (forward) - BUSCAR EN SubgrupoCursoLectivo
        curso_lectivo_id = self.forwarded.get('curso_lectivo', None)
        seccion_id = self.forwarded.get('seccion', None)
        especialidad_id = self.forwarded.get('especialidad', None)
        print(f"📅 Curso lectivo ID: {curso_lectivo_id}")
        print(f"📍 Sección ID: {seccion_id}")
        print(f"🎓 Especialidad ID: {especialidad_id}")
        print(f"📅 Forwarded completo: {self.forwarded}")
        
        if curso_lectivo_id and seccion_id:
            try:
                from config_institucional.models import SubgrupoCursoLectivo
                
                # Verificar que el curso lectivo existe (ahora es global)
                curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
                print(f"✅ Curso lectivo encontrado: {curso_lectivo.nombre}")
                
                # Verificar que la sección existe
                seccion = Seccion.objects.get(id=seccion_id)
                print(f"✅ Sección encontrada: {seccion}")
                
                # Filtros base
                filtros = {
                    'institucion_id': institucion_id,
                    'curso_lectivo': curso_lectivo,
                    'activa': True,
                    'subgrupo__seccion': seccion
                }
                
                # SI HAY ESPECIALIDAD (niveles 10, 11, 12): filtrar por especialidad
                if especialidad_id:
                    print(f"🎓 Filtrando por ESPECIALIDAD: {especialidad_id}")
                    filtros['especialidad_curso_id'] = especialidad_id
                
                # Obtener subgrupos configurados y activos para este curso lectivo
                # que pertenezcan específicamente a la sección seleccionada
                # y opcionalmente a la especialidad seleccionada
                subgrupos_configurados = SubgrupoCursoLectivo.objects.filter(
                    **filtros
                ).values_list('subgrupo_id', flat=True)
                
                print(f"🎯 Subgrupos configurados IDs para sección {seccion}: {list(subgrupos_configurados)}")
                
                # Filtrar subgrupos
                qs = Subgrupo.objects.filter(id__in=subgrupos_configurados)
                print(f"📋 Subgrupos encontrados: {[f'{s.letra} - Sección {s.seccion.numero}' for s in qs]}")
                
            except (CursoLectivo.DoesNotExist, Seccion.DoesNotExist, ValueError) as e:
                print(f"❌ Error: {e}")
                return Subgrupo.objects.none()
        else:
            # Sin curso lectivo o sección, no mostrar subgrupos
            if not curso_lectivo_id:
                print("❌ No hay curso lectivo seleccionado")
            if not seccion_id:
                print("❌ No hay sección seleccionada")
            return Subgrupo.objects.none()
        
        # Filtro por búsqueda
        if self.q:
            qs = qs.filter(letra__icontains=self.q)
            print(f"🔍 Filtrado por búsqueda '{self.q}': {[s.letra for s in qs]}")
        
        final_qs = qs.order_by('seccion__nivel__numero', 'seccion__numero', 'letra')
        print(f"🎯 RESULTADO FINAL: {[f'{s.letra} - Sección {s.seccion.numero}' for s in final_qs]}")
        return final_qs


# ════════════════════════════════════════════════════════════════
#                    ASIGNACIÓN AUTOMÁTICA DE GRUPOS
# ════════════════════════════════════════════════════════════════

@login_required
@permission_required('matricula.access_asignacion_grupos', raise_exception=True)
def asignacion_grupos(request):
    """
    Vista principal para la asignación automática de grupos.
    """
    from catalogos.models import CursoLectivo, Nivel
    from core.models import Institucion
    from .models import AsignacionGrupos
    
    context = {
        'instituciones': Institucion.objects.all().order_by('nombre') if request.user.is_superuser else [],
        'cursos_lectivos': CursoLectivo.objects.all().order_by('-anio'),
        'niveles': Nivel.objects.all().order_by('numero'),
        'es_superusuario': request.user.is_superuser,
        'asignaciones_recientes': AsignacionGrupos.objects.filter(
            institucion_id=getattr(request, 'institucion_activa_id', None) if not request.user.is_superuser else None
        ).order_by('-fecha_asignacion')[:10]
    }
    
    return render(request, 'matricula/asignacion_grupos.html', context)


@login_required
@permission_required('matricula.access_asignacion_grupos', raise_exception=True)
def ejecutar_asignacion_grupos(request):
    """
    Vista AJAX que ejecuta el algoritmo de asignación automática de grupos.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        # Obtener parámetros
        curso_lectivo_id = request.POST.get('curso_lectivo_id')
        nivel_id = request.POST.get('nivel_id')
        simular = request.POST.get('simular', 'false').lower() == 'true'
        
        # Validar parámetros requeridos
        if not curso_lectivo_id:
            return JsonResponse({'success': False, 'error': 'Curso lectivo es requerido'})
        
        # Obtener institución
        if request.user.is_superuser:
            institucion_id = request.POST.get('institucion_id')
            if not institucion_id:
                return JsonResponse({'success': False, 'error': 'Institución es requerida para superusuarios'})
            institucion = Institucion.objects.get(id=institucion_id)
        else:
            institucion_id = getattr(request, 'institucion_activa_id', None)
            if not institucion_id:
                return JsonResponse({'success': False, 'error': 'No se pudo determinar la institución activa'})
            institucion = Institucion.objects.get(id=institucion_id)
        
        # Obtener objetos
        from catalogos.models import CursoLectivo, Nivel
        curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
        nivel = Nivel.objects.get(id=nivel_id) if nivel_id else None
        
        # Ejecutar algoritmo
        from .asignacion_algoritmo import ejecutar_asignacion_completa
        resultado = ejecutar_asignacion_completa(
            institucion=institucion,
            curso_lectivo=curso_lectivo,
            nivel=nivel,
            usuario=request.user,
            simular=simular
        )
        # Normalizar respuesta de error para el frontend
        if not resultado.get('success'):
            mensaje_error = (
                resultado.get('mensaje')
                or (resultado.get('errores')[0] if resultado.get('errores') else None)
                or 'Error desconocido'
            )
            resultado['error'] = mensaje_error
        return JsonResponse(resultado)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en ejecutar_asignacion_grupos: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'})


@login_required
def exportar_listas_clase_excel(request):
    """
    Exporta listas de clase a Excel con filtros opcionales:
    - alcance: all | nivel | seccion | subgrupo
    - curso_lectivo_id (requerido)
    - nivel_id | seccion_id | subgrupo_id (dependiendo de alcance)
    """
    if openpyxl is None:
        return HttpResponse('openpyxl no está instalado en el entorno', status=500)

    alcance = request.GET.get('alcance', 'all')
    curso_lectivo_id = request.GET.get('curso_lectivo_id')
    nivel_id = request.GET.get('nivel_id')
    seccion_id = request.GET.get('seccion_id')
    subgrupo_id = request.GET.get('subgrupo_id')

    if not curso_lectivo_id:
        return HttpResponse('Debe indicar curso_lectivo_id', status=400)

    try:
        curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
    except CursoLectivo.DoesNotExist:
        return HttpResponse('Curso lectivo no encontrado', status=404)

    # Base queryset restringido por institución si no es superusuario
    qs = MatriculaAcademica.objects.select_related(
        'estudiante', 'nivel', 'seccion', 'subgrupo', 'especialidad__especialidad'
    ).filter(curso_lectivo=curso_lectivo, estado__iexact='activo')
    if not request.user.is_superuser:
        institucion_id = getattr(request, 'institucion_activa_id', None)
        if not institucion_id:
            return HttpResponse('No se pudo determinar la institución activa', status=400)
        qs = qs.filter(institucion_id=institucion_id)

    titulo = f"Listas de clase - {curso_lectivo.nombre}"

    if alcance == 'nivel' and nivel_id:
        qs = qs.filter(nivel_id=nivel_id)
    elif alcance == 'seccion' and seccion_id:
        qs = qs.filter(seccion_id=seccion_id)
    elif alcance == 'subgrupo' and subgrupo_id:
        qs = qs.filter(subgrupo_id=subgrupo_id)

    # Orden consistente alfabético por apellidos y nombres
    qs = qs.order_by('nivel__numero', 'seccion__numero', 'subgrupo__letra', 'estudiante__primer_apellido', 'estudiante__segundo_apellido', 'estudiante__nombres')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Listas'

    headers = [
        'Institución', 'Nivel', 'Sección', 'Subgrupo', 'Identificación',
        '1er Apellido', '2do Apellido', 'Nombres', 'Sexo', 'Especialidad'
    ]
    ws.append(headers)

    for mat in qs:
        ws.append([
            smart_str(getattr(mat.institucion, 'nombre', '')),  # Usar mat.institucion directamente
            smart_str(getattr(mat.nivel, 'nombre', '')),
            smart_str(f"{getattr(getattr(mat, 'nivel', None), 'numero', '')}-{getattr(getattr(mat, 'seccion', None), 'numero', '')}" if getattr(mat, 'seccion_id', None) and getattr(mat, 'nivel_id', None) else ''),
            smart_str(getattr(mat.subgrupo, 'letra', '')),
            smart_str(mat.estudiante.identificacion),
            smart_str(mat.estudiante.primer_apellido),
            smart_str(mat.estudiante.segundo_apellido),
            smart_str(mat.estudiante.nombres),
            smart_str(getattr(getattr(mat.estudiante, 'sexo', None), 'nombre', '')),
            smart_str(getattr(getattr(mat.especialidad, 'especialidad', None), 'nombre', '')),
        ])

    # Auto-ancho simple
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"listas_clase_{curso_lectivo.anio}_{alcance}.xlsx"
    response['Content-Disposition'] = f"attachment; filename={filename}"
    return response


@login_required
def api_secciones_por_curso_nivel(request):
    """Devuelve secciones activas (SeccionCursoLectivo) para curso_lectivo y nivel dados."""
    curso_lectivo_id = request.GET.get('curso_lectivo_id')
    nivel_id = request.GET.get('nivel_id')
    if not curso_lectivo_id or not nivel_id:
        return JsonResponse({'success': False, 'error': 'Parámetros incompletos'}, status=400)
    try:
        from config_institucional.models import SeccionCursoLectivo
        curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
        qs = SeccionCursoLectivo.objects.select_related('seccion', 'seccion__nivel').filter(
            curso_lectivo=curso_lectivo,
            activa=True,
            seccion__nivel_id=nivel_id
        )
        if not request.user.is_superuser:
            institucion_id = getattr(request, 'institucion_activa_id', None)
            if not institucion_id:
                return JsonResponse({'success': False, 'error': 'No se pudo determinar la institución'}, status=400)
            qs = qs.filter(institucion_id=institucion_id)
        data = [{'id': sc.seccion.id, 'nombre': f"{sc.seccion.nivel.numero}-{sc.seccion.numero}", 'numero': sc.seccion.numero} for sc in qs]
        return JsonResponse({'success': True, 'secciones': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def api_subgrupos_por_curso_seccion(request):
    """Devuelve subgrupos activos (SubgrupoCursoLectivo) para curso_lectivo y seccion dados."""
    curso_lectivo_id = request.GET.get('curso_lectivo_id')
    seccion_id = request.GET.get('seccion_id')
    if not curso_lectivo_id or not seccion_id:
        return JsonResponse({'success': False, 'error': 'Parámetros incompletos'}, status=400)
    try:
        from config_institucional.models import SubgrupoCursoLectivo
        curso_lectivo = CursoLectivo.objects.get(id=curso_lectivo_id)
        qs = SubgrupoCursoLectivo.objects.select_related('subgrupo', 'subgrupo__seccion').filter(
            curso_lectivo=curso_lectivo,
            activa=True,
            subgrupo__seccion_id=seccion_id
        )
        if not request.user.is_superuser:
            institucion_id = getattr(request, 'institucion_activa_id', None)
            if not institucion_id:
                return JsonResponse({'success': False, 'error': 'No se pudo determinar la institución'}, status=400)
            qs = qs.filter(institucion_id=institucion_id)
        data = [{'id': sc.subgrupo.id, 'nombre': f"{sc.subgrupo.seccion.nivel.numero}-{sc.subgrupo.seccion.numero}{sc.subgrupo.letra}", 'letra': sc.subgrupo.letra} for sc in qs]
        return JsonResponse({'success': True, 'subgrupos': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def buscar_estudiante_existente(request):
    """
    Busca si un estudiante con la identificación proporcionada ya existe en el sistema.
    Retorna información del estudiante y su institución activa.
    """
    identificacion = request.GET.get('identificacion', '').strip().upper()
    
    if not identificacion:
        return JsonResponse({'existe': False})
    
    try:
        # Buscar estudiante por identificación
        estudiante = Estudiante.objects.filter(identificacion=identificacion).first()
        
        if not estudiante:
            return JsonResponse({'existe': False})
        
        # Obtener la institución activa
        institucion_activa = estudiante.get_institucion_activa()
        
        # Verificar si el usuario es superusuario o tiene una institución
        institucion_usuario = None
        if not request.user.is_superuser:
            institucion_id = getattr(request, 'institucion_activa_id', None)
            if institucion_id:
                from core.models import Institucion
                institucion_usuario = Institucion.objects.filter(id=institucion_id).first()
        
        # Verificar si el estudiante ya pertenece a la institución del usuario
        ya_esta_en_institucion = False
        if institucion_usuario and institucion_activa:
            ya_esta_en_institucion = (institucion_activa.id == institucion_usuario.id)
        
        # Preparar datos del estudiante
        data = {
            'existe': True,
            'estudiante': {
                'id': estudiante.id,
                'identificacion': estudiante.identificacion,
                'nombres': estudiante.nombres,
                'primer_apellido': estudiante.primer_apellido,
                'segundo_apellido': estudiante.segundo_apellido,
                'nombre_completo': str(estudiante),
                'tipo_identificacion': estudiante.tipo_identificacion.nombre if estudiante.tipo_identificacion else '',
                'fecha_nacimiento': estudiante.fecha_nacimiento.strftime('%Y-%m-%d') if estudiante.fecha_nacimiento else '',
                'sexo': estudiante.sexo.id if estudiante.sexo else None,
                'nacionalidad': estudiante.nacionalidad.id if estudiante.nacionalidad else None,
                'correo': estudiante.correo,
                'celular': estudiante.celular or '',
                'telefono_casa': estudiante.telefono_casa or '',
                'provincia': estudiante.provincia.id if estudiante.provincia else None,
                'canton': estudiante.canton.id if estudiante.canton else None,
                'distrito': estudiante.distrito.id if estudiante.distrito else None,
                'direccion_exacta': estudiante.direccion_exacta or '',
            },
            'institucion_activa': {
                'id': institucion_activa.id if institucion_activa else None,
                'nombre': institucion_activa.nombre if institucion_activa else 'Sin institución activa'
            },
            'ya_esta_en_institucion': ya_esta_en_institucion
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al buscar estudiante: {e}")
        return JsonResponse({'existe': False, 'error': str(e)}, status=500)


@login_required
def agregar_estudiante_a_institucion(request):
    """
    Agrega un estudiante existente a la institución del usuario actual.
    Crea la relación EstudianteInstitucion con estado activo.
    Si el estudiante ya está activo en otra institución, marca esa relación como 'trasladado'.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        estudiante_id = data.get('estudiante_id')
        
        if not estudiante_id:
            return JsonResponse({'success': False, 'error': 'ID de estudiante no proporcionado'}, status=400)
        
        # Obtener el estudiante
        from matricula.models import Estudiante, EstudianteInstitucion
        estudiante = Estudiante.objects.filter(id=estudiante_id).first()
        
        if not estudiante:
            return JsonResponse({'success': False, 'error': 'Estudiante no encontrado'}, status=404)
        
        # Obtener la institución del usuario
        institucion_id = getattr(request, 'institucion_activa_id', None)
        if not institucion_id:
            return JsonResponse({'success': False, 'error': 'No se pudo determinar la institución del usuario'}, status=400)
        
        from core.models import Institucion
        institucion = Institucion.objects.filter(id=institucion_id).first()
        
        if not institucion:
            return JsonResponse({'success': False, 'error': 'Institución no encontrada'}, status=404)
        
        # Verificar si el estudiante está activo en otra institución
        relacion_activa_otra = EstudianteInstitucion.objects.filter(
            estudiante=estudiante,
            estado='activo'
        ).exclude(institucion=institucion).first()
        
        if relacion_activa_otra:
            return JsonResponse({
                'success': False,
                'error': f'El estudiante está activo en {relacion_activa_otra.institucion.nombre}. Esa institución debe darle de baja primero antes de poder agregarlo a su institución.',
                'requiere_baja': True,
                'institucion_actual': relacion_activa_otra.institucion.nombre
            })
        
        # Verificar si ya está activo en esta institución
        relacion_activa_actual = EstudianteInstitucion.objects.filter(
            estudiante=estudiante,
            institucion=institucion,
            estado='activo'
        ).first()
        
        if relacion_activa_actual:
            return JsonResponse({
                'success': False,
                'error': 'El estudiante ya está activo en esta institución'
            })
        
        # SIEMPRE crear un NUEVO registro para mantener el historial completo
        # No reutilizar registros anteriores aunque existieran
        fecha_actual = timezone.now().date()
        EstudianteInstitucion.objects.create(
            estudiante=estudiante,
            institucion=institucion,
            estado='activo',
            fecha_ingreso=fecha_actual,
            usuario_registro=request.user,
            observaciones=f'Ingreso a la institución el {fecha_actual.strftime("%d/%m/%Y")} - Agregado por {request.user.full_name() or request.user.email}'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Estudiante agregado exitosamente a la institución',
            'estudiante_id': estudiante.id
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al agregar estudiante a institución: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



