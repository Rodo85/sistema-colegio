from datetime import date
from decimal import Decimal
import csv
import hashlib
import logging
import re
import unicodedata

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError
from django.db import connection, transaction
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from catalogos.models import CursoLectivo, Nacionalidad, Sexo, TipoIdentificacion
from evaluaciones.models import (
    CentroTrabajo,
    DocenteAsignacion,
    EsquemaEvalComponente,
    Periodo,
    PeriodoCursoLectivo,
    SubareaCursoLectivo,
)
from config_institucional.models import Profesor
from matricula.models import Estudiante, EstudianteInstitucion, MatriculaAcademica, PlantillaImpresionMatricula

from .forms import (
    ActividadEvaluacionForm,
    AsignacionEditForm,
    AsignacionOnboardingForm,
    EstudianteCargaManualForm,
    IndicadorActividadFormSet,
)
from .models import ActividadEvaluacion, AsistenciaRegistro, AsistenciaSesion
from .models import PuntajeIndicador
from .models import ObservacionActividadEstudiante
from .models import PuntajeSimple
from .models import HorarioDocenteBloque, HorarioDocenteConfiguracion
from .models import EstudianteOcultoAsignacion
from .models import EstudianteAdecuacionAsignacion
from .models import EstudianteAdecuacionNoSignificativaAsignacion
from .models import ListaEstudiantesDocente, ListaEstudiantesDocenteItem
from .services import (
    actividad_pertenece_a_institucion,
    calcular_resumen_evaluacion_completo,
    calcular_resumen_componente_estudiante,
    calcular_total_maximo_actividad,
    copiar_actividad_a_asignaciones,
    duplicar_actividad,
    guardar_puntajes_masivo,
    obtener_porcentaje_componente_esquema,
    porcentaje_disponible_para_tipo,
    puede_usuario_editar_actividad,
)

try:
    import openpyxl
except Exception:
    openpyxl = None

# ─── Tabla: % ausencias injustificadas → asignación final (0-5) ──────────────
# Rangos: [min_inclusive, max_exclusive) → asignación
# 0% a <10% => 5, 10% a <20% => 4, 20% a <30% => 3, 30% a <40% => 2,
# 40% a <50% => 1, 50% o más => 0
_MEP_RANGES = [
    (0, 10, 5),
    (10, 20, 4),
    (20, 30, 3),
    (30, 40, 2),
    (40, 50, 1),
    (50, 100.01, 0),  # >= 50%
]

logger = logging.getLogger(__name__)

TIPOS_EVALUACION = (
    ActividadEvaluacion.TAREA,
    ActividadEvaluacion.COTIDIANO,
    ActividadEvaluacion.PRUEBA,
    ActividadEvaluacion.PROYECTO,
)


def _nota_mep(pct: float) -> int:
    """Convierte % ausencias injustificadas a asignación final 0-5."""
    pct = round(pct, 4)
    for min_pct, max_pct, nota in _MEP_RANGES:
        if min_pct <= pct < max_pct:
            return nota
    return 0


def _normalizar_estado_asistencia(estado):
    if estado == "T":
        return AsistenciaRegistro.TARDIA_MEDIA
    estados_validos = {k for k, _ in AsistenciaRegistro.ESTADO_CHOICES}
    return estado if estado in estados_validos else AsistenciaRegistro.PRESENTE


def _resolver_lecciones_injustificadas(
    estado,
    lecciones_dia,
    cantidad_ingresada=None,
    legacy_full_day_ai=True,
    cantidad_es_equivalente=False,
):
    """
    Convierte estado + cantidad a lecciones injustificadas equivalentes.

    Regla oficial por lección:
    - TM = 0.5 lección injustificada por unidad.
    - TC = 1 lección injustificada por unidad.
    - AI = 1 lección injustificada por unidad.
    - AJ/P = 0.

    Ejemplos:
    - Día de 6 lecciones y TM=2  -> 1.0 AI equivalente.
    - Día de 4 lecciones y AI=1  -> 1.0 AI equivalente.
    """
    estado = _normalizar_estado_asistencia(estado)
    lecciones = Decimal(str(lecciones_dia or 1))
    if lecciones < 1:
        lecciones = Decimal("1")

    if estado in (AsistenciaRegistro.PRESENTE, AsistenciaRegistro.AUSENTE_JUSTIFICADA):
        return Decimal("0")

    if cantidad_ingresada is not None:
        cantidad = Decimal(str(cantidad_ingresada))
        if cantidad < 0:
            raise ValidationError("La cantidad no puede ser negativa.")
        if cantidad > lecciones:
            raise ValidationError(f"La cantidad no puede exceder {lecciones}.")
        if cantidad_es_equivalente:
            # En registros persistidos, lecciones_injustificadas ya está en
            # equivalentes injustificadas (no en unidades TM).
            return cantidad
        if estado in (AsistenciaRegistro.TARDIA_MEDIA, AsistenciaRegistro.TARDIA_COMPLETA):
            if cantidad != cantidad.to_integral_value():
                raise ValidationError("Para TM y TC, la cantidad debe ser entera.")
        elif (cantidad * 2) != (cantidad * 2).to_integral_value():
            raise ValidationError("Para AI, la cantidad debe usar pasos de 0.5.")

        if estado == AsistenciaRegistro.TARDIA_MEDIA:
            return cantidad / Decimal("2")
        return cantidad

    # Compatibilidad de datos legacy sin cantidad explícita:
    if estado == AsistenciaRegistro.TARDIA_MEDIA:
        return min(Decimal("0.5"), lecciones)
    if estado == AsistenciaRegistro.TARDIA_COMPLETA:
        return min(Decimal("1"), lecciones)
    if estado == AsistenciaRegistro.AUSENTE_INJUSTIFICADA:
        return lecciones if legacy_full_day_ai else min(Decimal("1"), lecciones)
    return Decimal("0")


def _calcular_porcentajes_asistencia(total_lecciones, lecciones_injustificadas):
    total = Decimal(str(total_lecciones or 0))
    if total <= 0:
        return 0.0, 0.0
    inj = Decimal(str(lecciones_injustificadas or 0))
    if inj < 0:
        inj = Decimal("0")
    if inj > total:
        inj = total
    pct_inasistencia = float((inj / total) * Decimal("100"))
    pct_asistencia = max(0.0, 100.0 - pct_inasistencia)
    return pct_inasistencia, pct_asistencia


def _calcular_detalle_dia_asistencia(
    estado,
    lecciones_dia,
    cantidad_ingresada=None,
    legacy_full_day_ai=True,
    cantidad_es_equivalente=False,
):
    """
    Calcula detalle por día con dos salidas separadas:
    - Cantidades capturadas por estado (TM/TC/AI/AJ)
    - Lecciones injustificadas equivalentes (para % y asignación)
    """
    estado_norm = _normalizar_estado_asistencia(estado)
    lecciones = Decimal(str(lecciones_dia or 1))
    lecc_inj_equiv = _resolver_lecciones_injustificadas(
        estado=estado_norm,
        lecciones_dia=lecciones,
        cantidad_ingresada=cantidad_ingresada,
        legacy_full_day_ai=legacy_full_day_ai,
        cantidad_es_equivalente=cantidad_es_equivalente,
    )
    tm = Decimal("0")
    tc = Decimal("0")
    ai = Decimal("0")
    aj = Decimal("0")
    if estado_norm == AsistenciaRegistro.TARDIA_MEDIA:
        tm = lecc_inj_equiv * Decimal("2")
    elif estado_norm == AsistenciaRegistro.TARDIA_COMPLETA:
        tc = lecc_inj_equiv
    elif estado_norm == AsistenciaRegistro.AUSENTE_INJUSTIFICADA:
        ai = lecc_inj_equiv
    elif estado_norm == AsistenciaRegistro.AUSENTE_JUSTIFICADA:
        # AJ se visualiza en cantidad de lecciones justificadas.
        aj = lecciones

    presentes = max(Decimal("0"), lecciones - lecc_inj_equiv)
    return {
        "estado": estado_norm,
        "tm_cantidad": tm,
        "tc_cantidad": tc,
        "ai_cantidad": ai,
        "aj_cantidad": aj,
        "lecc_inj_equiv": lecc_inj_equiv,
        "presentes": presentes,
    }


def _get_profesor(request):
    """Devuelve el primer Profesor del usuario según la institución activa."""
    qs = Profesor.objects.filter(usuario=request.user).select_related("usuario")
    inst_id = getattr(request, "institucion_activa_id", None)
    if inst_id:
        qs = qs.filter(institucion_id=inst_id)
    return qs.first()


def _limite_asignaciones_docente(profesor):
    if not profesor:
        return None
    institucion = profesor.institucion
    if not getattr(institucion, "es_institucion_general", False):
        return None
    if profesor.max_asignaciones_override is not None:
        return profesor.max_asignaciones_override
    return institucion.max_asignaciones_general or 10


def _es_institucion_general(asignacion):
    try:
        return bool(asignacion.subarea_curso.institucion.es_institucion_general)
    except Exception:
        return False


def _es_institucion_general_profesor(profesor):
    return bool(profesor and getattr(profesor.institucion, "es_institucion_general", False))


def _asegurar_centro_principal(profesor):
    if not _es_institucion_general_profesor(profesor):
        return None
    centro, _ = CentroTrabajo.objects.get_or_create(
        docente=profesor,
        institucion=profesor.institucion,
        nombre="Centro principal",
        defaults={"activo": True},
    )
    if not centro.activo:
        centro.activo = True
        centro.save(update_fields=["activo", "updated_at"])
    return centro


def _es_centro_principal(centro):
    return bool(centro and (centro.nombre or "").strip().lower() == "centro principal")


def _obtener_lista_privada_docente(asignacion):
    filtros = {
        "docente": asignacion.docente,
        "institucion": asignacion.subarea_curso.institucion,
        "curso_lectivo": asignacion.curso_lectivo,
    }
    if asignacion.subgrupo_id:
        filtros["subgrupo_id"] = asignacion.subgrupo_id
    else:
        filtros["seccion_id"] = asignacion.seccion_id
    return ListaEstudiantesDocente.objects.filter(**filtros).first()


def _obtener_o_crear_lista_privada_docente(asignacion, user):
    filtros = {
        "docente": asignacion.docente,
        "institucion": asignacion.subarea_curso.institucion,
        "curso_lectivo": asignacion.curso_lectivo,
    }
    if asignacion.subgrupo_id:
        filtros["subgrupo_id"] = asignacion.subgrupo_id
    else:
        filtros["seccion_id"] = asignacion.seccion_id
    return ListaEstudiantesDocente.objects.get_or_create(
        defaults={"created_by": user},
        **filtros,
    )


def _obtener_estudiante_defaults():
    tipo_id = TipoIdentificacion.objects.order_by("id").first()
    sexo = Sexo.objects.order_by("id").first()
    nacionalidad = Nacionalidad.objects.order_by("id").first()
    if not tipo_id or not sexo or not nacionalidad:
        return None
    return {"tipo_identificacion": tipo_id, "sexo": sexo, "nacionalidad": nacionalidad}


def _normalizar_identificacion(valor):
    return re.sub(r"[\s-]+", "", str(valor or "").strip().upper())


def _es_tipo_cedula(tipo_identificacion):
    nombre = (getattr(tipo_identificacion, "nombre", "") or "").upper()
    return "CÉDULA" in nombre or "CEDULA" in nombre


def _colores_por_materia(nombre_materia):
    """
    Devuelve un par de colores consistente por nombre de materia.
    """
    paleta = [
        ("#3f7fb9", "#6fa8dc"),
        ("#5a73b8", "#7d94d1"),
        ("#4f8a8b", "#6db1b3"),
        ("#6f8f5a", "#8fb87a"),
        ("#7b6aa6", "#9a88c5"),
        ("#5f86a1", "#7ca5c3"),
        ("#6d8f94", "#8eb0b4"),
        ("#6379a0", "#8198bf"),
    ]
    base = (nombre_materia or "").strip().upper()
    digest = hashlib.md5(base.encode("utf-8")).hexdigest()
    idx = int(digest[:8], 16) % len(paleta)
    return paleta[idx]


def _dia_label_iso(dia_iso):
    labels = {
        1: "Lunes",
        2: "Martes",
        3: "Miércoles",
        4: "Jueves",
        5: "Viernes",
        6: "Sábado",
        7: "Domingo",
    }
    return labels.get(dia_iso, "Día")


def _periodo_id_para_asignacion(asignacion, fecha_ref=None):
    fecha_ref = fecha_ref or timezone.localdate()
    periodo = _infer_periodo(asignacion, fecha_ref)
    if periodo:
        return periodo.id
    first = (
        PeriodoCursoLectivo.objects.filter(
            institucion_id=asignacion.subarea_curso.institucion_id,
            curso_lectivo=asignacion.curso_lectivo,
            activo=True,
        )
        .order_by("periodo__numero")
        .values_list("periodo_id", flat=True)
        .first()
    )
    return first


def _acciones_rapidas_asignacion(asignacion, fecha_ref=None):
    periodo_id = _periodo_id_para_asignacion(asignacion, fecha_ref=fecha_ref)
    fecha_qs = ""
    if fecha_ref:
        try:
            fecha_qs = f"?fecha={fecha_ref.isoformat()}"
        except Exception:
            fecha_qs = ""
    tipos = _tipos_habilitados_por_esquema(asignacion)
    acciones = {
        "asistencia": reverse("libro_docente:asistencia", args=[asignacion.id]) + fecha_qs,
        "estudiantes": reverse("libro_docente:estudiantes_config", args=[asignacion.id]),
        "minuta": reverse("libro_docente:asistencia", args=[asignacion.id]) + fecha_qs,
        "detalle": reverse("libro_docente:resumen_evaluacion", args=[asignacion.id]),
        "tarea": None,
        "prueba": None,
    }
    if periodo_id:
        if ActividadEvaluacion.TAREA in tipos:
            acciones["tarea"] = (
                reverse("libro_docente:actividad_create", args=[asignacion.id])
                + f"?periodo={periodo_id}&tipo={ActividadEvaluacion.TAREA}"
            )
        if ActividadEvaluacion.PRUEBA in tipos:
            acciones["prueba"] = (
                reverse("libro_docente:actividad_create", args=[asignacion.id])
                + f"?periodo={periodo_id}&tipo={ActividadEvaluacion.PRUEBA}"
            )
    return acciones


def _parse_recesos_config(config):
    raw = (getattr(config, "receso_despues_leccion", "") or "").strip()
    if not raw:
        return []
    out = []
    for p in raw.split(","):
        p = p.strip()
        if not p.isdigit():
            continue
        n = int(p)
        if 1 <= n < config.max_lecciones_dia:
            out.append(n)
    return sorted(set(out))


def _color_por_clave(clave):
    paleta_ui = ["#4F46E5", "#F59E0B", "#06B6D4", "#10B981", "#EC4899"]
    base = (clave or "").strip().upper()
    if not base:
        return (paleta_ui[0], paleta_ui[0])
    digest = hashlib.md5(base.encode("utf-8")).hexdigest()
    idx = int(digest[:8], 16) % len(paleta_ui)
    color = paleta_ui[idx]
    return (color, color)


def _color_y_etiqueta_horario(asignacion):
    if asignacion.subgrupo_id:
        key = f"SUBGRUPO:{asignacion.subgrupo_id}"
        etiqueta = f"{asignacion.subgrupo} · {asignacion.subarea_curso.subarea.nombre}"
    else:
        key = f"MATERIA:{asignacion.subarea_curso.subarea.nombre}"
        sec = str(asignacion.seccion) if asignacion.seccion_id else "—"
        etiqueta = f"{sec} · {asignacion.subarea_curso.subarea.nombre}"
    color_primario, color_secundario = _color_por_clave(key)
    return key, etiqueta, color_primario, color_secundario


def _nombre_corto_materia(nombre):
    txt = (nombre or "").strip()
    if not txt:
        return "MATERIA"
    palabras = [p for p in txt.replace("-", " ").split() if p]
    stop = {"DE", "DEL", "LA", "LAS", "EL", "LOS", "Y", "E", "EN", "PARA"}
    utiles = [p for p in palabras if p.upper() not in stop]
    base = utiles or palabras
    if len(base) >= 2:
        sigla = "".join(p[0] for p in base[:4]).upper()
        if len(sigla) >= 2:
            return sigla
    compacto = "".join(ch for ch in txt.upper() if ch.isalnum())
    return compacto[:8] if compacto else "MATERIA"


def _nombre_corto_asignacion(asignacion):
    alias = (getattr(asignacion, "nombre_corto", "") or "").strip().upper()
    if alias:
        return alias
    return _nombre_corto_materia(asignacion.subarea_curso.subarea.nombre)


def _lecciones_programadas_para_fecha(asignacion, fecha_ref):
    """
    Cuenta lecciones del horario para una asignación en una fecha dada.
    Suma todos los bloques del día para esa asignación exacta.
    """
    if not asignacion or not fecha_ref:
        return 0
    try:
        dia_iso = int(fecha_ref.isoweekday())
    except Exception:
        logger.warning(
            "asistencia_lecciones: fecha inválida asignacion=%s fecha_ref=%s",
            getattr(asignacion, "id", None),
            fecha_ref,
        )
        return 0
    total_directo = (
        HorarioDocenteBloque.objects.filter(
            docente_asignacion=asignacion,
            dia_semana=dia_iso,
        )
        .values("leccion_numero")
        .distinct()
        .count()
    )
    if total_directo > 0:
        logger.info(
            "asistencia_lecciones: directo asignacion=%s fecha=%s dia=%s total=%s",
            asignacion.id,
            fecha_ref,
            dia_iso,
            total_directo,
        )
        return total_directo

    # Fallback defensivo: si el horario quedó asociado a otra asignación
    # equivalente (mismo docente/materia/grupo/centro), sumar esas lecciones.
    filtros = Q(
        docente_asignacion__docente_id=asignacion.docente_id,
        docente_asignacion__curso_lectivo_id=asignacion.curso_lectivo_id,
        docente_asignacion__subarea_curso__subarea_id=asignacion.subarea_curso.subarea_id,
        dia_semana=dia_iso,
    )
    if asignacion.subgrupo_id:
        filtros &= Q(docente_asignacion__subgrupo_id=asignacion.subgrupo_id)
    else:
        filtros &= Q(
            docente_asignacion__seccion_id=asignacion.seccion_id,
            docente_asignacion__subgrupo__isnull=True,
        )
    if asignacion.centro_trabajo_id:
        filtros &= Q(docente_asignacion__centro_trabajo_id=asignacion.centro_trabajo_id)
    else:
        filtros &= Q(docente_asignacion__centro_trabajo__isnull=True)
    total_fallback = (
        HorarioDocenteBloque.objects.filter(filtros)
        .values("leccion_numero")
        .distinct()
        .count()
    )
    resumen_directo = list(
        HorarioDocenteBloque.objects.filter(docente_asignacion=asignacion)
        .values("dia_semana")
        .annotate(total=Count("leccion_numero", distinct=True))
        .order_by("dia_semana")
    )
    logger.warning(
        "asistencia_lecciones: fallback asignacion=%s fecha=%s dia=%s total_directo=%s total_fallback=%s resumen_directo=%s",
        asignacion.id,
        fecha_ref,
        dia_iso,
        total_directo,
        total_fallback,
        resumen_directo,
    )
    return total_fallback


def _formatear_cantidad_asistencia(valor):
    if valor is None:
        return None
    try:
        dec = Decimal(str(valor))
    except Exception:
        return str(valor)
    dec = dec.quantize(Decimal("0.1")) if (dec * 10) == (dec * 10).to_integral_value() else dec
    texto = format(dec, "f")
    if "." in texto:
        texto = texto.rstrip("0").rstrip(".")
    return texto or "0"


def _limpiar_exclusiones_legacy(docente_asignacion_id):
    """
    Compatibilidad con despliegues que aún conservan tabla legacy de exclusiones.
    Si existe, limpiamos los registros para no bloquear el borrado de la asignación.
    """
    tabla_legacy = "libro_exclusion_estudiante_asignacion"
    columnas = ("docente_asignacion_id", "asignacion_id")
    if tabla_legacy not in connection.introspection.table_names():
        return
    with connection.cursor() as cursor:
        for columna in columnas:
            try:
                cursor.execute(
                    f'DELETE FROM "{tabla_legacy}" WHERE "{columna}" = %s',
                    [docente_asignacion_id],
                )
                return
            except Exception:
                continue


def _leer_estudiantes_desde_archivo(archivo):
    nombre = (getattr(archivo, "name", "") or "").lower()
    filas = []
    if nombre.endswith(".csv"):
        contenido = archivo.read().decode("utf-8", errors="ignore").splitlines()
        reader = csv.DictReader(contenido)
        for row in reader:
            filas.append(row)
    else:
        if openpyxl is None:
            raise ValidationError("No se puede procesar Excel porque openpyxl no está disponible.")
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = {}
            for i, h in enumerate(headers):
                row[h] = r[i] if i < len(r) else None
            filas.append(row)
    return filas


def _normalizar_filas_estudiantes(filas):
    mapping = {
        "identificacion": {
            "identificacion",
            "identificacion id",
            "id",
            "cedula",
        },
        "primer_apellido": {
            "primer apellido",
            "primerapellido",
            "1er apellido",
            "apellido1",
            "apellido 1",
        },
        "segundo_apellido": {
            "segundo apellido",
            "segundoapellido",
            "2do apellido",
            "apellido2",
            "apellido 2",
        },
        "nombres": {
            "nombre",
            "nombres",
        },
    }

    def normalize_header(text):
        text = str(text or "").strip().lower()
        text = "".join(
            ch for ch in unicodedata.normalize("NFD", text)
            if unicodedata.category(ch) != "Mn"
        )
        text = re.sub(r"[^a-z0-9]+", " ", text)
        return re.sub(r"\s+", " ", text).strip()

    def pick(row, keys):
        for k in row.keys():
            k_norm = normalize_header(k)
            if k_norm in keys:
                return str(row.get(k) or "").strip()
        return ""

    normalizadas = []
    for row in filas:
        ident = pick(row, mapping["identificacion"])
        p1 = pick(row, mapping["primer_apellido"])
        p2 = pick(row, mapping["segundo_apellido"])
        nom = pick(row, mapping["nombres"])
        if not ident and not p1 and not p2 and not nom:
            continue
        normalizadas.append(
            {
                "identificacion": ident,
                "primer_apellido": p1,
                "segundo_apellido": p2,
                "nombres": nom,
            }
        )
    return normalizadas


def _get_estudiantes(asignacion):
    """
    Devuelve MatriculaAcademica activas del grupo de la asignación,
    ordenadas por apellido.

    Regla: si hay subgrupo_id (materia técnica o asignación por subgrupo),
    filtrar SOLO por subgrupo. Si solo hay seccion_id (materia académica),
    filtrar por sección completa. Nunca mezclar 9-1A y 9-1B cuando
    la asignación es a un subgrupo específico.
    """
    filtros = {
        "curso_lectivo": asignacion.curso_lectivo,
        "estado": "activo",
        "institucion": asignacion.subarea_curso.institucion,
    }
    if asignacion.subgrupo_id:
        filtros["subgrupo_id"] = asignacion.subgrupo_id
    elif asignacion.seccion_id:
        filtros["seccion_id"] = asignacion.seccion_id
    else:
        return MatriculaAcademica.objects.none()
    ocultos_ids = list(
        EstudianteOcultoAsignacion.objects.filter(docente_asignacion=asignacion).values_list(
            "estudiante_id", flat=True
        )
    )
    qs = MatriculaAcademica.objects.filter(**filtros)
    if _es_institucion_general(asignacion):
        lista = _obtener_lista_privada_docente(asignacion)
        if not lista:
            return MatriculaAcademica.objects.none()
        est_ids = list(lista.items.values_list("estudiante_id", flat=True))
        if not est_ids:
            return MatriculaAcademica.objects.none()
        qs = qs.filter(estudiante_id__in=est_ids)
    if ocultos_ids:
        qs = qs.exclude(estudiante_id__in=ocultos_ids)
    return (
        qs.select_related("estudiante")
        .order_by("estudiante__primer_apellido", "estudiante__segundo_apellido", "estudiante__nombres")
    )


def _get_estudiantes_base(asignacion):
    """
    Lista base oficial del grupo/subgrupo sin aplicar ocultos.
    """
    filtros = {
        "curso_lectivo": asignacion.curso_lectivo,
        "estado": "activo",
        "institucion": asignacion.subarea_curso.institucion,
    }
    if asignacion.subgrupo_id:
        filtros["subgrupo_id"] = asignacion.subgrupo_id
    elif asignacion.seccion_id:
        filtros["seccion_id"] = asignacion.seccion_id
    else:
        return MatriculaAcademica.objects.none()
    qs = MatriculaAcademica.objects.filter(**filtros)
    if _es_institucion_general(asignacion):
        lista = _obtener_lista_privada_docente(asignacion)
        if not lista:
            return MatriculaAcademica.objects.none()
        qs = qs.filter(estudiante_id__in=lista.items.values_list("estudiante_id", flat=True))
    return (
        qs
        .select_related("estudiante")
        .order_by("estudiante__primer_apellido", "estudiante__segundo_apellido", "estudiante__nombres")
    )


def _get_ids_adecuacion(asignacion):
    return set(
        EstudianteAdecuacionAsignacion.objects.filter(docente_asignacion=asignacion).values_list(
            "estudiante_id", flat=True
        )
    )


def _get_ids_adecuacion_no_significativa(asignacion):
    return set(
        EstudianteAdecuacionNoSignificativaAsignacion.objects.filter(
            docente_asignacion=asignacion
        ).values_list("estudiante_id", flat=True)
    )


def _get_ids_adecuacion_reporte(asignacion):
    return _get_ids_adecuacion(asignacion).union(_get_ids_adecuacion_no_significativa(asignacion))


def _get_estudiantes_para_actividad(asignacion, actividad):
    """
    Filtra estudiantes según alcance de la actividad:
    - GRUPO: excluye estudiantes marcados con adecuación.
    - ADECUACION: muestra solo estudiantes marcados con adecuación.
    """
    qs = _get_estudiantes(asignacion)
    ids_adecuacion = _get_ids_adecuacion(asignacion)
    alcance = actividad.alcance_estudiantes
    if alcance == "GRUPO":
        alcance = ActividadEvaluacion.ALCANCE_REGULARES
    if alcance == ActividadEvaluacion.ALCANCE_TODOS:
        return qs
    if not ids_adecuacion:
        return qs if alcance != ActividadEvaluacion.ALCANCE_ADECUACION else qs.none()
    if alcance == ActividadEvaluacion.ALCANCE_ADECUACION:
        return qs.filter(estudiante_id__in=ids_adecuacion)
    return qs.exclude(estudiante_id__in=ids_adecuacion)


def _tipos_habilitados_por_esquema(asignacion):
    """
    Tipos visibles según componentes del esquema de evaluación snapshot.
    """
    if not asignacion.eval_scheme_snapshot_id:
        return []
    tipos = []
    componentes = (
        EsquemaEvalComponente.objects
        .filter(esquema=asignacion.eval_scheme_snapshot)
        .select_related("componente")
    )
    for c in componentes:
        cod = (c.componente.codigo or "").strip().upper()
        if cod in ("TAR", "TAREAS", "TAREA") and ActividadEvaluacion.TAREA not in tipos:
            tipos.append(ActividadEvaluacion.TAREA)
        elif cod in ("COT", "COTIDIANO") and ActividadEvaluacion.COTIDIANO not in tipos:
            tipos.append(ActividadEvaluacion.COTIDIANO)
        elif cod in ("PRU", "PRUEBA", "PRUEBAS") and ActividadEvaluacion.PRUEBA not in tipos:
            tipos.append(ActividadEvaluacion.PRUEBA)
        elif cod in ("PRO", "PROYECTO", "PROYECTOS") and ActividadEvaluacion.PROYECTO not in tipos:
            tipos.append(ActividadEvaluacion.PROYECTO)
    return tipos


def _infer_periodo(asignacion, fecha):
    """
    Infiere el Periodo a partir de la fecha usando PeriodoCursoLectivo.
    Devuelve el Periodo si hay coincidencia exacta por fechas;
    si no hay coincidencia, devuelve None.
    """
    inst_id = asignacion.subarea_curso.institucion_id
    periodos_cl = (
        PeriodoCursoLectivo.objects
        .filter(
            institucion_id=inst_id,
            curso_lectivo=asignacion.curso_lectivo,
        )
        .select_related("periodo")
        .order_by("periodo__numero")
    )
    for pcl in periodos_cl:
        if pcl.fecha_inicio and pcl.fecha_fin:
            if pcl.fecha_inicio <= fecha <= pcl.fecha_fin:
                return pcl.periodo
    return None


def _sesiones_por_periodo(asignacion, periodo):
    """
    Obtiene sesiones de una asignación para un período, priorizando rango de fechas
    del PeriodoCursoLectivo para incluir registros retroactivos válidos.
    """
    qs = AsistenciaSesion.objects.filter(docente_asignacion=asignacion)
    if not periodo:
        return qs.none()

    pcl = (
        PeriodoCursoLectivo.objects
        .filter(
            institucion_id=asignacion.subarea_curso.institucion_id,
            curso_lectivo=asignacion.curso_lectivo,
            periodo=periodo,
        )
        .first()
    )
    if pcl and pcl.fecha_inicio and pcl.fecha_fin:
        return qs.filter(fecha__range=(pcl.fecha_inicio, pcl.fecha_fin)).order_by("fecha", "sesion_numero")
    return qs.filter(periodo=periodo).order_by("fecha", "sesion_numero")


def _calcular_resumen(asignacion, periodo, matriculas):
    """
    Calcula resumen de asistencia por estudiante en un período.
    Regla vigente: el período se calcula por lecciones, no por cantidad
    de sesiones.
    """
    sesiones = _sesiones_por_periodo(asignacion, periodo)
    sesiones = list(sesiones)
    total_sesiones = len(sesiones)
    sesion_ids = [s.id for s in sesiones]
    matricula_est_ids = [m.estudiante_id for m in matriculas]
    registros_raw = AsistenciaRegistro.objects.filter(
        sesion_id__in=sesion_ids,
        estudiante_id__in=matricula_est_ids,
    )
    registros_map = {
        (r.estudiante_id, r.sesion_id): r
        for r in registros_raw
    }

    # Peso del componente ASISTENCIA en el esquema snapshot.
    # Se busca por código exacto "ASISTENCIA" o "ASIS", y también por
    # nombre que contenga "asistencia" — sin distinción de mayúsculas.
    peso_asistencia = Decimal("0")
    comp_asistencia = None
    if asignacion.eval_scheme_snapshot_id:
        comp_asistencia = (
            EsquemaEvalComponente.objects
            .filter(esquema=asignacion.eval_scheme_snapshot)
            .filter(
                Q(componente__codigo__iregex=r"^asis(tencia)?$") |
                Q(componente__nombre__icontains="asistencia")
            )
            .select_related("componente")
            .first()
        )
        if comp_asistencia:
            peso_asistencia = comp_asistencia.porcentaje

    resultados = []
    for m in matriculas:
        est = m.estudiante
        sesiones_est = list(sesiones)
        total_lecciones = sum((s.lecciones or 1) for s in sesiones_est)

        presentes = Decimal("0")
        tardias_media = Decimal("0")
        tardias_completa = Decimal("0")
        ausentes_ai = Decimal("0")
        ausentes_inj = Decimal("0")
        ausentes_just = Decimal("0")
        for s in sesiones_est:
            lecciones = Decimal(str(s.lecciones or 1))
            reg = registros_map.get((est.id, s.id))
            if reg is None:
                ausentes_inj += lecciones
                ausentes_ai += lecciones
                continue
            estado = _normalizar_estado_asistencia(reg.estado)
            try:
                detalle = _calcular_detalle_dia_asistencia(
                    estado=estado,
                    lecciones_dia=lecciones,
                    cantidad_ingresada=reg.lecciones_injustificadas,
                    legacy_full_day_ai=True,
                    cantidad_es_equivalente=True,
                )
            except ValidationError:
                detalle = _calcular_detalle_dia_asistencia(
                    estado=estado,
                    lecciones_dia=lecciones,
                    cantidad_ingresada=None,
                    legacy_full_day_ai=True,
                )
            ausentes_inj += detalle["lecc_inj_equiv"]
            presentes += detalle["presentes"]
            tardias_media += detalle["tm_cantidad"]
            tardias_completa += detalle["tc_cantidad"]
            # AI y AJ deben reflejar la cantidad real capturada/interpretada en su estado.
            # AI ya no mezcla TM/TC.
            ai_cantidad = detalle["ai_cantidad"]
            aj_cantidad = detalle["aj_cantidad"]
            if ai_cantidad:
                ausentes_ai += ai_cantidad
            if aj_cantidad:
                ausentes_just += aj_cantidad

        pct, pct_asistencia = _calcular_porcentajes_asistencia(total_lecciones, ausentes_inj)
        puntaje_base = _nota_mep(pct)
        # aporte_real = (asignacion_final / 5) * peso_asistencia_esquema
        aporte_real = (
            Decimal(str(puntaje_base)) / Decimal("5") * peso_asistencia
            if peso_asistencia else Decimal("0")
        )

        # Indicador visual por cumplimiento de asistencia >= 80%
        if total_lecciones == 0:
            nivel_alerta = "nodata"
        elif pct_asistencia >= 80:
            nivel_alerta = "ok"
        else:
            nivel_alerta = "danger"

        resultados.append({
            "estudiante": est,
            "presentes": presentes.quantize(Decimal("0.1")),
            "tardias_media": tardias_media.quantize(Decimal("0.1")),
            "tardias_completa": tardias_completa.quantize(Decimal("0.1")),
            # Cantidad real marcada de AI (sin mezclar equivalencias de TM/TC)
            "ausentes_inj_lecciones": ausentes_ai.quantize(Decimal("0.1")),
            # Total equivalente injustificado (se usa para % de ausencia)
            "ausentes_inj_equiv": ausentes_inj.quantize(Decimal("0.01")),
            "ausentes_just": ausentes_just.quantize(Decimal("0.1")),
            "total_lecciones": total_lecciones,
            "pct": round(pct, 2),
            "pct_asistencia": round(pct_asistencia, 2),
            "nota_mep": puntaje_base,
            "peso_asistencia": peso_asistencia,
            "aporte_real": round(aporte_real, 2),
            "nivel_alerta": nivel_alerta,
        })

    nombre_componente = (
        comp_asistencia.componente.nombre if comp_asistencia else "Asistencia"
    )
    resultados.sort(
        key=lambda r: (
            r["estudiante"].primer_apellido or "",
            r["estudiante"].segundo_apellido or "",
            r["estudiante"].nombres or "",
        )
    )
    return {
        "total_sesiones": total_sesiones,
        "total_lecciones_periodo": sum((s.lecciones or 1) for s in sesiones),
        "peso_asistencia": peso_asistencia,
        "tiene_componente": comp_asistencia is not None,
        "nombre_componente": nombre_componente,
        "estudiantes": resultados,
    }


# ═══════════════════════════════════════════════════════════════════════
#  VISTAS
# ═══════════════════════════════════════════════════════════════════════

@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def home_docente(request):
    """Home del docente: tarjetas por cada DocenteAsignacion activa."""
    profesor = _get_profesor(request)
    error = None
    asignaciones_data = []
    limite_asignaciones = None

    centros_trabajo = []
    materias_filtro = []
    niveles_filtro = []
    grupos_filtro = []
    clases_hoy = []
    centro_sel_id = 0
    materia_sel_id = 0
    nivel_sel_id = 0
    grupo_sel = ""
    if not profesor:
        error = "No tienes perfil de docente registrado en esta institución."
    else:
        es_general = _es_institucion_general_profesor(profesor)
        if es_general:
            _asegurar_centro_principal(profesor)
            centros_trabajo = list(
                CentroTrabajo.objects.filter(
                    docente=profesor,
                    institucion=profesor.institucion,
                    activo=True,
                ).order_by("nombre")
            )
            try:
                centro_sel_id = int(request.GET.get("centro", "0") or "0")
            except (TypeError, ValueError):
                centro_sel_id = 0
            if not request.session.get("msg_centros_trabajo_general_v1"):
                messages.info(
                    request,
                    "Ahora puedes organizar tus asignaciones por Centro de trabajo. "
                    "Tus datos actuales fueron ubicados en 'Centro principal'.",
                )
                request.session["msg_centros_trabajo_general_v1"] = True

        try:
            materia_sel_id = int(request.GET.get("materia", "0") or "0")
        except (TypeError, ValueError):
            materia_sel_id = 0
        try:
            nivel_sel_id = int(request.GET.get("nivel", "0") or "0")
        except (TypeError, ValueError):
            nivel_sel_id = 0
        grupo_sel = (request.GET.get("grupo", "") or "").strip()

        limite_asignaciones = _limite_asignaciones_docente(profesor)
        dia_hoy_iso = timezone.localdate().isoweekday()
        clases_hoy_qs = (
            HorarioDocenteBloque.objects.filter(
                configuracion__docente=profesor,
                configuracion__institucion=profesor.institucion,
                dia_semana=dia_hoy_iso,
                docente_asignacion__activo=True,
            )
            .select_related(
                "configuracion__centro_trabajo",
                "docente_asignacion__subarea_curso__subarea",
                "docente_asignacion__seccion__nivel",
                "docente_asignacion__subgrupo__seccion__nivel",
            )
            .order_by("leccion_numero")
        )
        if es_general and centro_sel_id:
            clases_hoy_qs = clases_hoy_qs.filter(configuracion__centro_trabajo_id=centro_sel_id)
        for b in clases_hoy_qs:
            asig = b.docente_asignacion
            grupo_label = str(asig.subgrupo) if asig.subgrupo_id else (str(asig.seccion) if asig.seccion_id else "—")
            color_primario, color_secundario = _colores_por_materia(asig.subarea_curso.subarea.nombre)
            clases_hoy.append({
                "leccion": b.leccion_numero,
                "asignacion": asig,
                "materia": asig.subarea_curso.subarea.nombre,
                "grupo_label": grupo_label,
                "centro": b.configuracion.centro_trabajo.nombre if b.configuracion.centro_trabajo_id else "",
                "color_primario": color_primario,
                "color_secundario": color_secundario,
                "acciones": _acciones_rapidas_asignacion(asig),
            })

        componentes_prefetch = Prefetch(
            "eval_scheme_snapshot__componentes_esquema",
            queryset=EsquemaEvalComponente.objects.select_related("componente").order_by("componente__nombre"),
        )
        raw_qs = (
            DocenteAsignacion.objects
            .filter(docente=profesor, activo=True)
            .select_related(
                "subarea_curso__subarea",
                "curso_lectivo",
                "seccion__nivel",
                "subgrupo__seccion__nivel",
                "centro_trabajo",
                "eval_scheme_snapshot",
            )
            .prefetch_related(componentes_prefetch)
        )
        if es_general and centro_sel_id:
            raw_qs = raw_qs.filter(centro_trabajo_id=centro_sel_id)

        # Opciones de filtro (sobre el universo ya acotado por centro)
        materias_filtro = list(
            raw_qs.values("subarea_curso__subarea_id", "subarea_curso__subarea__nombre")
            .distinct()
            .order_by("subarea_curso__subarea__nombre")
        )
        niveles_map = {}
        grupos_map = {}
        for a in raw_qs:
            if a.subgrupo_id:
                nivel_id = a.subgrupo.seccion.nivel_id
                nivel_numero = a.subgrupo.seccion.nivel.numero
                nivel_label = str(nivel_numero)
                grupo_value = f"SUB-{a.subgrupo_id}"
                grupo_label = str(a.subgrupo)
            elif a.seccion_id:
                nivel_id = a.seccion.nivel_id
                nivel_numero = a.seccion.nivel.numero
                nivel_label = str(nivel_numero)
                grupo_value = f"SEC-{a.seccion_id}"
                grupo_label = str(a.seccion)
            else:
                continue
            niveles_map[nivel_id] = nivel_label
            grupos_map[grupo_value] = grupo_label
        niveles_filtro = [
            {"id": nid, "label": niveles_map[nid]}
            for nid in sorted(niveles_map, key=lambda x: int(niveles_map[x]) if str(niveles_map[x]).isdigit() else 999)
        ]
        grupos_filtro = [
            {"value": gval, "label": glabel}
            for gval, glabel in sorted(grupos_map.items(), key=lambda x: x[1])
        ]

        # Aplicar filtros seleccionados
        if materia_sel_id:
            raw_qs = raw_qs.filter(subarea_curso__subarea_id=materia_sel_id)
        if nivel_sel_id:
            raw_qs = raw_qs.filter(
                Q(seccion__nivel_id=nivel_sel_id) | Q(subgrupo__seccion__nivel_id=nivel_sel_id)
            )
        if grupo_sel.startswith("SEC-"):
            try:
                raw_qs = raw_qs.filter(seccion_id=int(grupo_sel.split("-", 1)[1]))
            except (TypeError, ValueError):
                pass
        elif grupo_sel.startswith("SUB-"):
            try:
                raw_qs = raw_qs.filter(subgrupo_id=int(grupo_sel.split("-", 1)[1]))
            except (TypeError, ValueError):
                pass
        raw = list(raw_qs)

        def _sort_key(a):
            if a.subgrupo_id:
                n = a.subgrupo.seccion.nivel.numero
                s = a.subgrupo.seccion.numero
                l = (a.subgrupo.letra or "").upper()
                return (n, s, l)
            if a.seccion_id:
                n = a.seccion.nivel.numero
                s = a.seccion.numero
                return (n, s, "")
            return (999, 999, "")

        raw.sort(key=lambda a: (_sort_key(a), a.subarea_curso.subarea.nombre))
        materias_distintas = {
            (a.subarea_curso.subarea.nombre or "").strip().upper()
            for a in raw
            if a.subarea_curso_id and a.subarea_curso.subarea_id
        }
        paleta_ui = ["#4F46E5", "#F59E0B", "#06B6D4", "#10B981", "#EC4899"]
        if materias_distintas:
            materias_ordenadas = sorted(materias_distintas)
            paleta_activa = paleta_ui[: min(len(paleta_ui), len(materias_ordenadas))]
            color_por_materia = {
                materia: paleta_activa[idx % len(paleta_activa)]
                for idx, materia in enumerate(materias_ordenadas)
            }
            color_default = paleta_activa[0]
        else:
            color_por_materia = {}
            color_default = paleta_ui[0]

        # Sesiones hoy por asignación (1 query para todas)
        asignacion_ids = [a.id for a in raw]
        sesiones_hoy_counts = dict(
            AsistenciaSesion.objects.filter(
                docente_asignacion_id__in=asignacion_ids, fecha=timezone.localdate()
            ).values("docente_asignacion_id").annotate(cnt=Count("id")).values_list("docente_asignacion_id", "cnt")
        )

        hoy = timezone.localdate()
        for a in raw:
            componentes_raw = list(a.eval_scheme_snapshot.componentes_esquema.all()) if a.eval_scheme_snapshot_id else []
            componentes = []
            for c in componentes_raw:
                cod = (c.componente.codigo or "").strip().upper()
                if cod in ("TAR", "TAREAS", "TAREA"):
                    tipo_param = "TAREA"
                elif cod in ("COT", "COTIDIANO"):
                    tipo_param = "COTIDIANO"
                elif cod in ("PRU", "PRUEBA", "PRUEBAS"):
                    tipo_param = "PRUEBA"
                elif cod in ("PRO", "PROYECTO", "PROYECTOS"):
                    tipo_param = "PROYECTO"
                else:
                    tipo_param = None
                componentes.append({
                    "componente": c.componente,
                    "porcentaje": c.porcentaje,
                    "tipo_param": tipo_param,
                })

            tiene_asistencia = any(
                (c["componente"].codigo or "").upper() in ("ASISTENCIA", "ASIS") or
                "ASISTENCIA" in (c["componente"].nombre or "").upper()
                for c in componentes
            )

            sesiones_hoy = sesiones_hoy_counts.get(a.id, 0)

            # Etiqueta del grupo: subgrupo completo (7-1A) o sección (7-1)
            if a.subgrupo_id:
                grupo_label = str(a.subgrupo)  # ej. 7-1A, 8-3B
            elif a.seccion_id:
                grupo_label = str(a.seccion)   # ej. 7-1
            else:
                grupo_label = "—"
            materia_key = (a.subarea_curso.subarea.nombre or "").strip().upper()
            color = color_por_materia.get(materia_key, color_default)
            color_primario, color_secundario = (color, color)

            asignaciones_data.append({
                "obj": a,
                "componentes": componentes,
                "tiene_asistencia": tiene_asistencia,
                "sesiones_hoy": sesiones_hoy,
                "grupo_label": grupo_label,
                "centro_trabajo": a.centro_trabajo.nombre if a.centro_trabajo_id else "",
                "color_primario": color_primario,
                "color_secundario": color_secundario,
            })

    can_create_asignacion = bool(
        profesor and (
            limite_asignaciones is None or len(asignaciones_data) < limite_asignaciones
        )
    )

    return render(request, "libro_docente/hoy.html", {
        "asignaciones": asignaciones_data,
        "hoy": timezone.localdate(),
        "profesor": profesor,
        "error": error,
        "show_onboarding": bool(profesor and not asignaciones_data),
        "limite_asignaciones": limite_asignaciones,
        "can_create_asignacion": can_create_asignacion,
        "es_institucion_general": _es_institucion_general_profesor(profesor),
        "centros_trabajo": centros_trabajo,
        "centro_sel_id": centro_sel_id,
        "materias_filtro": materias_filtro,
        "niveles_filtro": niveles_filtro,
        "grupos_filtro": grupos_filtro,
        "materia_sel_id": materia_sel_id,
        "nivel_sel_id": nivel_sel_id,
        "grupo_sel": grupo_sel,
        "clases_hoy": clases_hoy,
        "dia_hoy_label": _dia_label_iso(timezone.localdate().isoweekday()),
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def asignacion_onboarding_view(request):
    profesor = _get_profesor(request)
    if not profesor:
        messages.error(request, "No tienes perfil docente para crear asignaciones.")
        return redirect("libro_docente:home")

    institucion = profesor.institucion
    if _es_institucion_general_profesor(profesor):
        _asegurar_centro_principal(profesor)
    curso_lectivo = CursoLectivo.get_activo()
    if not curso_lectivo:
        curso_lectivo = (
            CursoLectivo.objects
            .order_by("-anio", "-id")
            .first()
        )
    if not curso_lectivo:
        messages.error(request, "No hay curso lectivo disponible.")
        return redirect("libro_docente:home")

    form = AsignacionOnboardingForm(
        request.POST or None,
        institucion=institucion,
        curso_lectivo=curso_lectivo,
        profesor=profesor,
    )

    if request.method == "POST" and form.is_valid():
        subarea = form.cleaned_data["subarea"]
        esquema = form.cleaned_data["eval_scheme"]
        centro = form.cleaned_data.get("centro_trabajo")
        sec_cl = form.cleaned_data.get("seccion")
        sgr_cl = form.cleaned_data.get("subgrupo")

        scl, created = SubareaCursoLectivo.objects.get_or_create(
            institucion=institucion,
            curso_lectivo=curso_lectivo,
            subarea=subarea,
            defaults={"activa": True, "eval_scheme": esquema},
        )
        updates = []
        if not scl.activa:
            scl.activa = True
            updates.append("activa")
        if not scl.eval_scheme_id:
            scl.eval_scheme = esquema
            updates.append("eval_scheme")
        if updates:
            scl.save(update_fields=updates)

        asignacion = DocenteAsignacion(
            docente=profesor,
            subarea_curso=scl,
            curso_lectivo=curso_lectivo,
            seccion=sec_cl.seccion if sec_cl else None,
            subgrupo=sgr_cl.subgrupo if sgr_cl else None,
            centro_trabajo=centro,
            nombre_corto=(form.cleaned_data.get("nombre_corto") or ""),
            activo=True,
            eval_scheme_snapshot=esquema,
        )
        try:
            asignacion.full_clean()
            asignacion.save()
            messages.success(request, "Asignación creada correctamente. Ya puedes usar el Libro del Docente.")
            return redirect("libro_docente:home")
        except ValidationError as exc:
            form.add_error(None, exc)

    return render(request, "libro_docente/asignacion_onboarding.html", {
        "form": form,
        "curso_lectivo": curso_lectivo,
        "institucion": institucion,
        "limite_asignaciones": _limite_asignaciones_docente(profesor),
        "es_institucion_general": _es_institucion_general_profesor(profesor),
        "is_edit": False,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def centros_trabajo_view(request):
    profesor = _get_profesor(request)
    if not profesor:
        messages.error(request, "No tienes perfil docente registrado en esta institución.")
        return redirect("libro_docente:home")
    if not _es_institucion_general_profesor(profesor):
        messages.error(request, "Los centros de trabajo aplican solo para Institución General.")
        return redirect("libro_docente:home")

    _asegurar_centro_principal(profesor)
    centros = list(
        CentroTrabajo.objects.filter(
            docente=profesor,
            institucion=profesor.institucion,
        ).order_by("nombre")
    )
    centro_editar = None
    editar_id = request.GET.get("editar")
    if editar_id and str(editar_id).isdigit():
        centro_editar = CentroTrabajo.objects.filter(
            id=int(editar_id),
            docente=profesor,
            institucion=profesor.institucion,
        ).first()

    if request.method == "POST":
        accion = (request.POST.get("accion") or "crear").strip().lower()
        if accion == "mover_asignacion":
            asignacion_id = request.POST.get("asignacion_id")
            destino_id = request.POST.get("centro_destino_id")
            if not (asignacion_id and str(asignacion_id).isdigit() and destino_id and str(destino_id).isdigit()):
                messages.error(request, "Datos inválidos para mover la asignación.")
                return redirect("libro_docente:centros_trabajo")
            asignacion = get_object_or_404(
                DocenteAsignacion,
                id=int(asignacion_id),
                docente=profesor,
                activo=True,
            )
            destino = get_object_or_404(
                CentroTrabajo,
                id=int(destino_id),
                docente=profesor,
                institucion=profesor.institucion,
                activo=True,
            )
            if asignacion.centro_trabajo_id == destino.id:
                messages.info(request, "La asignación ya estaba en ese centro.")
                return redirect("libro_docente:centros_trabajo")
            asignacion.centro_trabajo = destino
            try:
                asignacion.save()
                messages.success(request, "Asignación movida correctamente.")
            except ValidationError as exc:
                messages.error(request, f"No se pudo mover la asignación: {exc}")
            return redirect("libro_docente:centros_trabajo")

        if accion == "eliminar":
            centro_id = request.POST.get("centro_id")
            if not (centro_id and str(centro_id).isdigit()):
                messages.error(request, "Centro de trabajo no válido.")
                return redirect("libro_docente:centros_trabajo")
            centro = get_object_or_404(
                CentroTrabajo,
                id=int(centro_id),
                docente=profesor,
                institucion=profesor.institucion,
            )
            if _es_centro_principal(centro):
                messages.error(request, "No se puede eliminar el Centro principal.")
                return redirect("libro_docente:centros_trabajo")
            total_asignaciones = DocenteAsignacion.objects.filter(
                docente=profesor,
                centro_trabajo=centro,
            ).count()
            if total_asignaciones > 0:
                messages.error(
                    request,
                    f"No se puede eliminar. Este centro tiene {total_asignaciones} asignación(es). "
                    "Muévelas primero a otro centro.",
                )
                return redirect("libro_docente:centros_trabajo")
            centro.delete()
            messages.success(request, "Centro de trabajo eliminado.")
            return redirect("libro_docente:centros_trabajo")

        nombre = (request.POST.get("nombre") or "").strip()
        logo = request.FILES.get("logo")
        if not nombre:
            messages.error(request, "Debes indicar el nombre del centro de trabajo.")
            return redirect("libro_docente:centros_trabajo")

        if accion == "editar":
            centro_id = request.POST.get("centro_id")
            if not (centro_id and str(centro_id).isdigit()):
                messages.error(request, "Centro de trabajo no válido.")
                return redirect("libro_docente:centros_trabajo")
            centro = get_object_or_404(
                CentroTrabajo,
                id=int(centro_id),
                docente=profesor,
                institucion=profesor.institucion,
            )
            if _es_centro_principal(centro):
                messages.error(request, "No se puede editar el Centro principal.")
                return redirect("libro_docente:centros_trabajo")
            centro.nombre = nombre
            if logo:
                centro.logo = logo
            centro.save()
            messages.success(request, "Centro de trabajo actualizado.")
            return redirect("libro_docente:centros_trabajo")

        _, created = CentroTrabajo.objects.get_or_create(
            docente=profesor,
            institucion=profesor.institucion,
            nombre=nombre,
            defaults={"activo": True, "logo": logo},
        )
        if created:
            messages.success(request, "Centro de trabajo creado.")
        else:
            messages.info(request, "Ese centro de trabajo ya existe.")
        return redirect("libro_docente:centros_trabajo")

    conteos = dict(
        DocenteAsignacion.objects.filter(
            docente=profesor,
            activo=True,
            centro_trabajo_id__in=[c.id for c in centros],
        ).values("centro_trabajo_id").annotate(total=Count("id")).values_list("centro_trabajo_id", "total")
    )
    centros_info = [
        {"obj": c, "asignaciones": conteos.get(c.id, 0), "es_principal": _es_centro_principal(c)}
        for c in centros
    ]
    asignaciones = list(
        DocenteAsignacion.objects.filter(docente=profesor, activo=True)
        .select_related("subarea_curso__subarea", "seccion__nivel", "subgrupo__seccion__nivel", "centro_trabajo")
        .order_by("centro_trabajo__nombre", "subarea_curso__subarea__nombre")
    )
    return render(request, "libro_docente/centros_trabajo.html", {
        "profesor": profesor,
        "centros": centros_info,
        "centro_editar": centro_editar,
        "asignaciones": asignaciones,
        "centros_destino": centros,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def horario_docente_view(request):
    profesor = _get_profesor(request)
    if not profesor:
        messages.error(request, "No tienes perfil docente registrado en esta institución.")
        return redirect("libro_docente:home")

    institucion = profesor.institucion
    es_general = _es_institucion_general_profesor(profesor)
    centros = []
    centro_sel_id = 0
    centro_actual = None
    if es_general:
        _asegurar_centro_principal(profesor)
        centros = list(
            CentroTrabajo.objects.filter(
                docente=profesor,
                institucion=institucion,
                activo=True,
            ).order_by("nombre")
        )
        try:
            centro_sel_id = int((request.POST.get("centro") if request.method == "POST" else request.GET.get("centro")) or 0)
        except (TypeError, ValueError):
            centro_sel_id = 0
        if not centro_sel_id and centros:
            centro_sel_id = centros[0].id
        centro_actual = next((c for c in centros if c.id == centro_sel_id), None)
        if not centro_actual and centros:
            centro_actual = centros[0]
            centro_sel_id = centro_actual.id
        if not centro_actual:
            centro_actual = _asegurar_centro_principal(profesor)
            if centro_actual:
                centro_sel_id = centro_actual.id

    config_lookup = {
        "docente": profesor,
        "institucion": institucion,
        "centro_trabajo": centro_actual if es_general else None,
    }
    config, _ = HorarioDocenteConfiguracion.objects.get_or_create(
        **config_lookup,
        defaults={"max_lecciones_dia": 8},
    )
    weekdays = [
        (HorarioDocenteBloque.LUNES, "Lunes"),
        (HorarioDocenteBloque.MARTES, "Martes"),
        (HorarioDocenteBloque.MIERCOLES, "Miércoles"),
        (HorarioDocenteBloque.JUEVES, "Jueves"),
        (HorarioDocenteBloque.VIERNES, "Viernes"),
    ]
    asignaciones_qs = (
        DocenteAsignacion.objects.filter(docente=profesor, activo=True)
        .select_related("subarea_curso__subarea", "seccion__nivel", "subgrupo__seccion__nivel", "centro_trabajo")
        .order_by("subarea_curso__subarea__nombre")
    )
    if es_general:
        asignaciones_qs = asignaciones_qs.filter(centro_trabajo=centro_actual)
    asignaciones = list(asignaciones_qs)
    asignaciones_ids = {a.id for a in asignaciones}

    if request.method == "POST" and request.POST.get("accion") == "guardar_horario":
        max_raw = request.POST.get("max_lecciones_dia")
        try:
            max_lecciones = int(max_raw)
        except (TypeError, ValueError):
            max_lecciones = config.max_lecciones_dia
        max_lecciones = max(1, min(20, max_lecciones))

        recesos_raw = request.POST.getlist("recesos_despues_leccion")
        recesos = []
        for r in recesos_raw:
            try:
                n = int(r)
            except (TypeError, ValueError):
                continue
            if 1 <= n < max_lecciones:
                recesos.append(n)
        recesos = sorted(set(recesos))
        recesos_csv = ",".join(str(n) for n in recesos)

        updates = []
        if config.max_lecciones_dia != max_lecciones:
            config.max_lecciones_dia = max_lecciones
            updates.append("max_lecciones_dia")
        if (config.receso_despues_leccion or "") != recesos_csv:
            config.receso_despues_leccion = recesos_csv
            updates.append("receso_despues_leccion")
        if updates:
            updates.append("updated_at")
            config.save(update_fields=updates)

        nuevos = []
        for dia, _ in weekdays:
            for lec in range(1, config.max_lecciones_dia + 1):
                k = f"h_{dia}_{lec}"
                asignacion_id = request.POST.get(k)
                if not (asignacion_id and str(asignacion_id).isdigit()):
                    continue
                asignacion_id = int(asignacion_id)
                if asignacion_id not in asignaciones_ids:
                    continue
                nuevos.append(
                    HorarioDocenteBloque(
                        configuracion=config,
                        dia_semana=dia,
                        leccion_numero=lec,
                        docente_asignacion_id=asignacion_id,
                    )
                )
        with transaction.atomic():
            config.bloques.all().delete()
            if nuevos:
                HorarioDocenteBloque.objects.bulk_create(nuevos)
        messages.success(request, "Horario guardado correctamente.")
        next_url = reverse("libro_docente:horario_docente")
        if es_general and centro_sel_id:
            next_url += f"?centro={centro_sel_id}"
        return redirect(next_url)

    bloques = list(
        HorarioDocenteBloque.objects.filter(configuracion=config)
        .select_related("docente_asignacion__subarea_curso__subarea")
    )
    recesos = _parse_recesos_config(config)
    celdas = {(b.dia_semana, b.leccion_numero): b.docente_asignacion_id for b in bloques}
    asig_by_id = {a.id: a for a in asignaciones}

    filas_edicion = []
    for lec in range(1, config.max_lecciones_dia + 1):
        celdas_fila = []
        for dia, _dia_label in weekdays:
            aid = celdas.get((dia, lec))
            celdas_fila.append({
                "dia": dia,
                "leccion": lec,
                "asignacion_id": aid or "",
            })
        filas_edicion.append({"leccion": lec, "celdas": celdas_fila})

    hidden_cells = set()
    merged_blocks = {}
    for dia, _ in weekdays:
        lec = 1
        while lec <= config.max_lecciones_dia:
            aid = celdas.get((dia, lec))
            if not aid:
                lec += 1
                continue
            span = 1
            nxt = lec + 1
            while nxt <= config.max_lecciones_dia and celdas.get((dia, nxt)) == aid:
                if (nxt - 1) in recesos:
                    break
                span += 1
                nxt += 1
            asignacion = asig_by_id.get(aid)
            if asignacion:
                grupo_label = str(asignacion.subgrupo) if asignacion.subgrupo_id else (str(asignacion.seccion) if asignacion.seccion_id else "—")
                _k, _e, color_primario, color_secundario = _color_y_etiqueta_horario(asignacion)
                merged_blocks[(dia, lec)] = {
                    "rowspan": span,
                    "asignacion": asignacion,
                    "grupo_label": grupo_label,
                    "materia_corta": _nombre_corto_asignacion(asignacion),
                    "materia_full": asignacion.subarea_curso.subarea.nombre,
                    "color_primario": color_primario,
                    "color_secundario": color_secundario,
                    "acciones": _acciones_rapidas_asignacion(asignacion),
                }
            for x in range(lec + 1, lec + span):
                hidden_cells.add((dia, x))
            lec += span

    row_defs = []
    for lec in range(1, config.max_lecciones_dia + 1):
        row_defs.append({"tipo": "leccion", "leccion": lec})
        if lec in recesos:
            row_defs.append({"tipo": "receso"})

    tabla_compacta = []
    for row in row_defs:
        if row["tipo"] == "receso":
            tabla_compacta.append({"tipo": "receso"})
            continue
        lec = row["leccion"]
        cells = []
        for dia, _ in weekdays:
            key = (dia, lec)
            if key in hidden_cells:
                cells.append({"skip": True, "block": None})
            else:
                cells.append({"skip": False, "block": merged_blocks.get(key)})
        tabla_compacta.append({"tipo": "leccion", "leccion": lec, "celdas": cells})

    day_headers = []
    for idx, (dia, dia_label) in enumerate(weekdays, start=1):
        day_headers.append({
            "dia": dia,
            "dia_label": dia_label,
            "col_start": idx + 1,  # Columna 1 es el índice de lección
        })

    grid_rows = []
    leccion_to_grid_row = {}
    for idx, row in enumerate(row_defs, start=1):
        entry = {"tipo": row["tipo"], "row_num": idx + 1}  # Fila 1 es encabezado
        if row["tipo"] == "leccion":
            entry["leccion"] = row["leccion"]
            leccion_to_grid_row[row["leccion"]] = idx + 1
        grid_rows.append(entry)

    grid_blocks = []
    for head in day_headers:
        dia = head["dia"]
        for lec in range(1, config.max_lecciones_dia + 1):
            block = merged_blocks.get((dia, lec))
            if not block:
                continue
            row_start = leccion_to_grid_row.get(lec)
            if not row_start:
                continue
            row_end = row_start + block["rowspan"]
            grid_blocks.append({
                "col_start": head["col_start"],
                "row_start": row_start,
                "row_end": row_end,
                "block": block,
            })

    receso_bands = []
    first_day_col = 2
    last_day_col = len(day_headers) + 2
    for row in grid_rows:
        if row["tipo"] == "receso":
            receso_bands.append({
                "row_num": row["row_num"],
                "col_start": first_day_col,
                "col_end": last_day_col,
            })
    row_tracks = ["38px"]
    for row in row_defs:
        if row["tipo"] == "receso":
            row_tracks.append("24px")
        else:
            row_tracks.append("54px")
    grid_template_rows = " ".join(row_tracks)

    return render(request, "libro_docente/horario_docente.html", {
        "profesor": profesor,
        "es_institucion_general": es_general,
        "centros_trabajo": centros,
        "centro_sel_id": centro_sel_id,
        "centro_actual": centro_actual,
        "config": config,
        "asignaciones": asignaciones,
        "filas_edicion": filas_edicion,
        "dias": weekdays,
        "tabla_compacta": tabla_compacta,
        "recesos_sel": recesos,
        "grid_day_headers": day_headers,
        "grid_rows": grid_rows,
        "grid_blocks": grid_blocks,
        "grid_receso_bands": receso_bands,
        "grid_total_cols": len(day_headers) + 1,
        "grid_total_rows": len(grid_rows) + 1,
        "grid_template_rows": grid_template_rows,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def asignacion_estudiantes_excel_view(request, asignacion_id):
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")
    if not _es_institucion_general(asignacion):
        messages.error(request, "La carga por Excel solo está disponible en Institución General.")
        return redirect("libro_docente:home")

    form_manual = EstudianteCargaManualForm()
    lista_actual = _obtener_lista_privada_docente(asignacion)

    if request.method == "POST" and request.POST.get("accion") == "actualizar_lista":
        estudiante_id = request.POST.get("estudiante_id")
        if not (estudiante_id and str(estudiante_id).isdigit()):
            messages.error(request, "No se pudo identificar al estudiante para actualizar.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))
        if not lista_actual:
            messages.error(request, "No existe una lista cargada para esta asignación.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        item = ListaEstudiantesDocenteItem.objects.filter(
            lista=lista_actual,
            estudiante_id=int(estudiante_id),
        ).select_related("estudiante").first()
        if not item:
            messages.error(request, "El estudiante no pertenece a tu lista actual.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        p1 = (request.POST.get("primer_apellido") or "").strip().upper()
        p2 = (request.POST.get("segundo_apellido") or "").strip().upper()
        nom = (request.POST.get("nombres") or "").strip().upper()
        if not p1 or not p2 or not nom:
            messages.error(request, "Para actualizar, completa primer apellido, segundo apellido y nombre.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        est = item.estudiante
        updates = []
        if est.primer_apellido != p1:
            est.primer_apellido = p1
            updates.append("primer_apellido")
        if est.segundo_apellido != p2:
            est.segundo_apellido = p2
            updates.append("segundo_apellido")
        if est.nombres != nom:
            est.nombres = nom
            updates.append("nombres")
        if updates:
            est.save(update_fields=updates)
            messages.success(request, "Estudiante actualizado correctamente.")
        else:
            messages.info(request, "No hubo cambios para actualizar.")
        return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

    if request.method == "POST" and request.POST.get("accion") == "eliminar_lista":
        estudiante_id = request.POST.get("estudiante_id")
        if not (estudiante_id and str(estudiante_id).isdigit()):
            messages.error(request, "No se pudo identificar al estudiante para quitar de la lista.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))
        if not lista_actual:
            messages.error(request, "No existe una lista cargada para esta asignación.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        eliminado, _ = ListaEstudiantesDocenteItem.objects.filter(
            lista=lista_actual,
            estudiante_id=int(estudiante_id),
        ).delete()
        if eliminado:
            messages.success(request, "Estudiante quitado de tu lista. No se eliminó del sistema.")
        else:
            messages.warning(request, "El estudiante ya no estaba en tu lista.")
        return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

    if request.method == "POST" and request.POST.get("accion") == "agregar_manual":
        limite = 25 if asignacion.subgrupo_id else 50
        defaults_catalogo = _obtener_estudiante_defaults()
        if not defaults_catalogo:
            messages.error(
                request,
                "Faltan catálogos mínimos (Sexo/Nacionalidad) para crear estudiantes.",
            )
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        form_manual = EstudianteCargaManualForm(request.POST)
        if not form_manual.is_valid():
            messages.error(request, "Revisa los datos del formulario manual.")
        else:
            with transaction.atomic():
                lista, _ = _obtener_o_crear_lista_privada_docente(asignacion, request.user)
                total_actual = lista.items.count()
                if total_actual >= limite:
                    messages.error(
                        request,
                        f"Se alcanzó el límite permitido para este grupo: {limite} estudiantes.",
                    )
                    return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

                ident = _normalizar_identificacion(form_manual.cleaned_data["identificacion"])
                p1 = form_manual.cleaned_data["primer_apellido"].strip().upper()
                p2 = form_manual.cleaned_data["segundo_apellido"].strip().upper()
                nom = form_manual.cleaned_data["nombres"].strip().upper()
                tipo_id = form_manual.cleaned_data["tipo_identificacion"]

                est = Estudiante.objects.filter(identificacion=ident).first()
                estudiante_reutilizado = bool(est)
                if est:
                    est.tipo_identificacion = tipo_id
                    est.primer_apellido = p1
                    est.segundo_apellido = p2
                    est.nombres = nom
                    if not est.fecha_nacimiento:
                        est.fecha_nacimiento = date(2000, 1, 1)
                    if not est.sexo_id:
                        est.sexo = defaults_catalogo["sexo"]
                    if not est.nacionalidad_id:
                        est.nacionalidad = defaults_catalogo["nacionalidad"]
                    est.save()
                else:
                    est = Estudiante.objects.create(
                        tipo_estudiante=Estudiante.PR,
                        tipo_identificacion=tipo_id,
                        identificacion=ident,
                        primer_apellido=p1,
                        segundo_apellido=p2,
                        nombres=nom,
                        fecha_nacimiento=date(2000, 1, 1),
                        sexo=defaults_catalogo["sexo"],
                        nacionalidad=defaults_catalogo["nacionalidad"],
                        correo=f"{ident.lower()}@est.mep.go.cr",
                    )

                rel_activa = EstudianteInstitucion.objects.filter(
                    estudiante=est,
                    estado=EstudianteInstitucion.ACTIVO,
                ).first()
                if rel_activa and rel_activa.institucion_id != asignacion.subarea_curso.institucion_id:
                    messages.error(
                        request,
                        f"ID {ident}: el estudiante está activo en otra institución y no puede agregarse aquí.",
                    )
                    return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))
                if not rel_activa:
                    EstudianteInstitucion.objects.create(
                        estudiante=est,
                        institucion=asignacion.subarea_curso.institucion,
                        estado=EstudianteInstitucion.ACTIVO,
                        fecha_ingreso=timezone.now().date(),
                        usuario_registro=request.user,
                    )

                MatriculaAcademica.objects.update_or_create(
                    estudiante=est,
                    curso_lectivo=asignacion.curso_lectivo,
                    defaults={
                        "institucion": asignacion.subarea_curso.institucion,
                        "nivel": (asignacion.subgrupo.seccion.nivel if asignacion.subgrupo_id else asignacion.seccion.nivel),
                        "seccion": (asignacion.subgrupo.seccion if asignacion.subgrupo_id else asignacion.seccion),
                        "subgrupo": (asignacion.subgrupo if asignacion.subgrupo_id else None),
                        "estado": MatriculaAcademica.ACTIVO,
                        "origen_carga": MatriculaAcademica.ORIGEN_GENERAL_EXCEL,
                    },
                )

                item, created = ListaEstudiantesDocenteItem.objects.get_or_create(
                    lista=lista,
                    estudiante=est,
                    defaults={"orden": total_actual + 1},
                )
                if not created and item.orden <= 0:
                    item.orden = total_actual + 1
                    item.save(update_fields=["orden"])

            if estudiante_reutilizado:
                messages.success(
                    request,
                    "Estudiante agregado a tu lista usando el registro existente (sin duplicar).",
                )
            else:
                messages.success(request, "Estudiante agregado correctamente.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

    if request.method == "POST" and request.POST.get("accion") == "plantilla":
        if openpyxl is None:
            messages.error(request, "No se pudo generar la plantilla Excel.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"
        ws.append(["Identificación (ID)", "Primer apellido", "Segundo apellido", "Nombre"])
        ws.append(["123456789", "APELLIDO1", "APELLIDO2", "NOMBRE"])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="plantilla_estudiantes_{asignacion.id}.xlsx"'
        wb.save(response)
        return response

    if request.method == "POST" and request.POST.get("accion") == "subir":
        archivo = request.FILES.get("archivo_excel")
        if not archivo:
            messages.error(request, "Debes seleccionar un archivo Excel o CSV.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        try:
            filas_raw = _leer_estudiantes_desde_archivo(archivo)
            filas = _normalizar_filas_estudiantes(filas_raw)
        except Exception as exc:
            messages.error(request, f"No se pudo leer el archivo: {exc}")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        if not filas:
            messages.error(request, "El archivo no contiene filas válidas.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        ids = [_normalizar_identificacion(f.get("identificacion")) for f in filas if f.get("identificacion")]
        if len(ids) != len(set(ids)):
            messages.error(request, "El archivo contiene IDs repetidos. Corrige e intenta de nuevo.")
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        limite = 25 if asignacion.subgrupo_id else 50
        if len(filas) > limite:
            messages.error(
                request,
                f"Se supera el límite permitido para este grupo: {limite} estudiantes.",
            )
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        defaults_catalogo = _obtener_estudiante_defaults()
        if not defaults_catalogo:
            messages.error(
                request,
                "Faltan catálogos mínimos (TipoIdentificación/Sexo/Nacionalidad) para crear estudiantes.",
            )
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        creados = 0
        actualizados = 0
        errores = []
        agregados_lista = 0
        ya_en_lista = 0
        try:
            with transaction.atomic():
                lista, _ = _obtener_o_crear_lista_privada_docente(asignacion, request.user)
                estudiantes_lista = []
                for idx, f in enumerate(filas, start=2):
                    ident = _normalizar_identificacion(f.get("identificacion"))
                    p1 = (f.get("primer_apellido") or "").strip().upper()
                    p2 = (f.get("segundo_apellido") or "").strip().upper()
                    nom = (f.get("nombres") or "").strip().upper()
                    if not ident or not p1 or not p2 or not nom:
                        errores.append(f"Fila {idx}: faltan datos obligatorios.")
                        continue
                    if _es_tipo_cedula(defaults_catalogo["tipo_identificacion"]):
                        if not ident.isdigit() or len(ident) != 9:
                            errores.append(
                                f"Fila {idx}: cédula inválida ({ident}). Debe tener 9 dígitos sin guiones."
                            )
                            continue

                    est = Estudiante.objects.filter(identificacion=ident).first()
                    if est:
                        cambio = False
                        if est.primer_apellido != p1:
                            est.primer_apellido = p1
                            cambio = True
                        if est.segundo_apellido != p2:
                            est.segundo_apellido = p2
                            cambio = True
                        if est.nombres != nom:
                            est.nombres = nom
                            cambio = True
                        if cambio:
                            est.save()
                            actualizados += 1
                    else:
                        est = Estudiante.objects.create(
                            tipo_estudiante=Estudiante.PR,
                            tipo_identificacion=defaults_catalogo["tipo_identificacion"],
                            identificacion=ident,
                            primer_apellido=p1,
                            segundo_apellido=p2,
                            nombres=nom,
                            fecha_nacimiento=date(2000, 1, 1),
                            sexo=defaults_catalogo["sexo"],
                            nacionalidad=defaults_catalogo["nacionalidad"],
                            correo=f"{ident.lower()}@est.mep.go.cr",
                        )
                        creados += 1

                    rel_activa = EstudianteInstitucion.objects.filter(
                        estudiante=est,
                        estado=EstudianteInstitucion.ACTIVO,
                    ).first()
                    if rel_activa and rel_activa.institucion_id != asignacion.subarea_curso.institucion_id:
                        errores.append(
                            f"ID {ident}: el estudiante está activo en otra institución y no puede importarse en General."
                        )
                        continue
                    if not rel_activa:
                        EstudianteInstitucion.objects.create(
                            estudiante=est,
                            institucion=asignacion.subarea_curso.institucion,
                            estado=EstudianteInstitucion.ACTIVO,
                            fecha_ingreso=timezone.now().date(),
                            usuario_registro=request.user,
                        )

                    MatriculaAcademica.objects.update_or_create(
                        estudiante=est,
                        curso_lectivo=asignacion.curso_lectivo,
                        defaults={
                            "institucion": asignacion.subarea_curso.institucion,
                            "nivel": (asignacion.subgrupo.seccion.nivel if asignacion.subgrupo_id else asignacion.seccion.nivel),
                            "seccion": (asignacion.subgrupo.seccion if asignacion.subgrupo_id else asignacion.seccion),
                            "subgrupo": (asignacion.subgrupo if asignacion.subgrupo_id else None),
                            "estado": MatriculaAcademica.ACTIVO,
                            "origen_carga": MatriculaAcademica.ORIGEN_GENERAL_EXCEL,
                        },
                    )
                    estudiantes_lista.append(est.id)

                if not estudiantes_lista and errores:
                    raise ValidationError("No se pudo cargar ningún estudiante válido.")

                existentes = set(
                    ListaEstudiantesDocenteItem.objects.filter(lista=lista).values_list("estudiante_id", flat=True)
                )
                ya_en_lista = sum(1 for est_id in estudiantes_lista if est_id in existentes)
                nuevos_items = []
                base_orden = ListaEstudiantesDocenteItem.objects.filter(lista=lista).count()
                for offset, est_id in enumerate(estudiantes_lista, start=1):
                    if est_id not in existentes:
                        nuevos_items.append(
                            ListaEstudiantesDocenteItem(
                                lista=lista,
                                estudiante_id=est_id,
                                orden=base_orden + offset,
                            )
                        )
                agregados_lista = len(nuevos_items)
                if nuevos_items:
                    ListaEstudiantesDocenteItem.objects.bulk_create(nuevos_items)
        except ValidationError as exc:
            messages.error(request, str(exc))
            return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

        if errores:
            messages.warning(
                request,
                (
                    f"Carga parcial: agregados a la lista {agregados_lista}, "
                    f"ya existentes en la lista {ya_en_lista}, "
                    f"creados {creados}, actualizados {actualizados}, "
                    f"no cargados {len(errores)}."
                ),
            )
            max_detalle = 10
            detalle = " | ".join(errores[:max_detalle])
            if len(errores) > max_detalle:
                detalle += f" | ... y {len(errores) - max_detalle} error(es) más."
            messages.error(request, f"Detalle de no cargados: {detalle}")
        else:
            messages.success(
                request,
                (
                    f"Carga completada: agregados a la lista {agregados_lista}, "
                    f"ya existentes en la lista {ya_en_lista}, "
                    f"creados {creados}, actualizados {actualizados}, "
                    f"procesados {len(filas)}."
                ),
            )
        return redirect(reverse("libro_docente:asignacion_estudiantes_excel", args=[asignacion.id]))

    lista = _obtener_lista_privada_docente(asignacion)
    estudiantes = []
    if lista:
        estudiantes = list(
            MatriculaAcademica.objects.filter(
                curso_lectivo=asignacion.curso_lectivo,
                institucion=asignacion.subarea_curso.institucion,
                estudiante_id__in=lista.items.values_list("estudiante_id", flat=True),
                estado=MatriculaAcademica.ACTIVO,
            )
            .select_related("estudiante")
            .order_by("estudiante__primer_apellido", "estudiante__segundo_apellido", "estudiante__nombres")
        )

    return render(request, "libro_docente/asignacion_estudiantes_excel.html", {
        "asignacion": asignacion,
        "estudiantes": estudiantes,
        "limite": (25 if asignacion.subgrupo_id else 50),
        "form_manual": form_manual,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def asignacion_edit_view(request, asignacion_id):
    profesor = _get_profesor(request)
    if not profesor:
        messages.error(request, "No tienes perfil docente para editar asignaciones.")
        return redirect("libro_docente:home")

    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")

    form = AsignacionEditForm(request.POST or None, asignacion=asignacion)

    if request.method == "POST" and form.is_valid():
        esquema = form.cleaned_data["eval_scheme"]
        asignacion.nombre_corto = (form.cleaned_data.get("nombre_corto") or "")
        asignacion.eval_scheme_snapshot = esquema
        try:
            asignacion.full_clean()
            asignacion.save()
            messages.success(request, "Asignación actualizada correctamente.")
            return redirect("libro_docente:home")
        except ValidationError as exc:
            form.add_error(None, exc)

    return render(request, "libro_docente/asignacion_edit.html", {
        "form": form,
        "asignacion": asignacion,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def asignacion_delete_view(request, asignacion_id):
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")

    if request.method == "POST":
        if request.POST.get("confirmar") != "SI":
            messages.error(request, "Debes confirmar para eliminar la asignación.")
            return redirect(reverse("libro_docente:asignacion_delete", args=[asignacion.id]))
        with transaction.atomic():
            _limpiar_exclusiones_legacy(asignacion.id)
            AsistenciaSesion.objects.filter(docente_asignacion=asignacion).delete()
            ActividadEvaluacion.objects.filter(docente_asignacion=asignacion).delete()
            EstudianteOcultoAsignacion.objects.filter(docente_asignacion=asignacion).delete()
            EstudianteAdecuacionAsignacion.objects.filter(docente_asignacion=asignacion).delete()
            EstudianteAdecuacionNoSignificativaAsignacion.objects.filter(docente_asignacion=asignacion).delete()
            asignacion.delete()
        messages.success(
            request,
            "Asignación eliminada. Se borraron los registros asociados del Libro del Docente, pero no los estudiantes.",
        )
        return redirect("libro_docente:home")

    return render(request, "libro_docente/asignacion_confirm_delete.html", {
        "asignacion": asignacion,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def estudiantes_config_view(request, asignacion_id):
    """
    Pantalla independiente para gestionar estudiantes ocultos y con adecuación.
    """
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")

    base = list(_get_estudiantes_base(asignacion))
    base_ids = [m.estudiante_id for m in base]
    ocultos_actuales = set(
        EstudianteOcultoAsignacion.objects.filter(
            docente_asignacion=asignacion,
            estudiante_id__in=base_ids,
        ).values_list("estudiante_id", flat=True)
    )
    adecuacion_actuales = _get_ids_adecuacion(asignacion).intersection(set(base_ids))
    adecuacion_no_sig_actuales = _get_ids_adecuacion_no_significativa(asignacion).intersection(set(base_ids))

    if request.method == "POST":
        oculto_ids = {int(x) for x in request.POST.getlist("oculto_ids") if str(x).isdigit()}
        adecuacion_ids = {int(x) for x in request.POST.getlist("adecuacion_ids") if str(x).isdigit()}
        adecuacion_no_sig_ids = {
            int(x) for x in request.POST.getlist("adecuacion_no_sig_ids") if str(x).isdigit()
        }
        oculto_ids = oculto_ids.intersection(set(base_ids))
        adecuacion_ids = adecuacion_ids.intersection(set(base_ids))
        adecuacion_no_sig_ids = adecuacion_no_sig_ids.intersection(set(base_ids))

        with transaction.atomic():
            EstudianteOcultoAsignacion.objects.filter(
                docente_asignacion=asignacion,
                estudiante_id__in=base_ids,
            ).exclude(estudiante_id__in=oculto_ids).delete()
            EstudianteAdecuacionAsignacion.objects.filter(
                docente_asignacion=asignacion,
                estudiante_id__in=base_ids,
            ).exclude(estudiante_id__in=adecuacion_ids).delete()
            EstudianteAdecuacionNoSignificativaAsignacion.objects.filter(
                docente_asignacion=asignacion,
                estudiante_id__in=base_ids,
            ).exclude(estudiante_id__in=adecuacion_no_sig_ids).delete()

            for est_id in oculto_ids - ocultos_actuales:
                EstudianteOcultoAsignacion.objects.get_or_create(
                    docente_asignacion=asignacion,
                    estudiante_id=est_id,
                    defaults={"created_by": request.user},
                )
            for est_id in adecuacion_ids - adecuacion_actuales:
                EstudianteAdecuacionAsignacion.objects.get_or_create(
                    docente_asignacion=asignacion,
                    estudiante_id=est_id,
                    defaults={"created_by": request.user},
                )
            for est_id in adecuacion_no_sig_ids - adecuacion_no_sig_actuales:
                EstudianteAdecuacionNoSignificativaAsignacion.objects.get_or_create(
                    docente_asignacion=asignacion,
                    estudiante_id=est_id,
                    defaults={"created_by": request.user},
                )
        messages.success(request, "Configuración de estudiantes guardada.")
        return redirect(reverse("libro_docente:estudiantes_config", args=[asignacion.id]))

    filas = []
    for m in base:
        est = m.estudiante
        filas.append({
            "id": est.id,
            "nombre": str(est),
            "identificacion": est.identificacion,
            "oculto": est.id in ocultos_actuales,
            "adecuacion": est.id in adecuacion_actuales,
            "adecuacion_no_sig": est.id in adecuacion_no_sig_actuales,
        })
    filas.sort(key=lambda x: x["nombre"])
    return render(request, "libro_docente/estudiantes_config.html", {
        "asignacion": asignacion,
        "filas": filas,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def prueba_lista_ejecucion_view(request, actividad_id):
    """
    Lista imprimible para ejecución de pruebas:
    - Regulares
    - Adecuación (significativa y no significativa)
    """
    actividad = get_object_or_404(
        ActividadEvaluacion.objects.select_related(
            "docente_asignacion__subarea_curso__subarea",
            "docente_asignacion__curso_lectivo",
            "docente_asignacion__subgrupo__seccion__nivel",
            "docente_asignacion__seccion__nivel",
            "periodo",
        ),
        id=actividad_id,
    )
    if actividad.tipo_componente != ActividadEvaluacion.PRUEBA:
        messages.error(request, "Esta lista solo está disponible para actividades de tipo Prueba.")
        return redirect(reverse("libro_docente:actividad_list", args=[actividad.docente_asignacion_id]))

    asignacion = _obtener_asignacion_con_permiso(request, actividad.docente_asignacion_id)
    if asignacion is None or asignacion.id != actividad.docente_asignacion_id:
        messages.error(request, "No tienes acceso a esta actividad.")
        return redirect("libro_docente:home")

    matriculas = list(_get_estudiantes(asignacion))
    adec_sig = _get_ids_adecuacion(asignacion)
    adec_no_sig = _get_ids_adecuacion_no_significativa(asignacion)
    adec_reporte = adec_sig.union(adec_no_sig)

    regulares = []
    adecuacion = []
    for m in matriculas:
        est = m.estudiante
        row = {
            "id": est.identificacion,
            "nombre": str(est),
            "tipo": (
                "Adecuación significativa"
                if est.id in adec_sig
                else "Adecuación no significativa"
                if est.id in adec_no_sig
                else ""
            ),
        }
        if est.id in adec_reporte:
            adecuacion.append(row)
        else:
            regulares.append(row)

    plantilla = PlantillaImpresionMatricula.objects.filter(
        institucion=asignacion.subarea_curso.institucion
    ).first()
    grupo_label = str(asignacion.subgrupo) if asignacion.subgrupo_id else str(asignacion.seccion)
    return render(request, "libro_docente/prueba_lista_ejecucion.html", {
        "actividad": actividad,
        "asignacion": asignacion,
        "grupo_label": grupo_label,
        "regulares": regulares,
        "adecuacion": adecuacion,
        "plantilla": plantilla,
        "total_estudiantes": len(regulares) + len(adecuacion),
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def asistencia_view(request, asignacion_id):
    """
    Pantalla de asistencia diaria: una sola sesión por fecha, con cantidad
    de lecciones del día y marcas rápidas por estudiante.
    """
    profesor = _get_profesor(request)
    if not profesor:
        messages.error(request, "No tienes perfil de docente en esta institución.")
        return redirect("libro_docente:home")

    asignacion = get_object_or_404(
        DocenteAsignacion.objects.select_related(
            "subarea_curso__subarea", "curso_lectivo",
            "seccion__nivel", "subgrupo__seccion__nivel", "centro_trabajo",
        ),
        id=asignacion_id, docente=profesor, activo=True,
    )

    # ── Fecha ────────────────────────────────────────────────────────────
    hoy = timezone.localdate()
    fecha_str = (request.POST if request.method == "POST" else request.GET).get("fecha")
    try:
        fecha = date.fromisoformat(fecha_str) if fecha_str else hoy
    except ValueError:
        fecha = hoy

    # ── Lecciones del día ────────────────────────────────────────────────
    lecciones_programadas = _lecciones_programadas_para_fecha(asignacion, fecha)
    raw_lecciones = (request.POST if request.method == "POST" else request.GET).get("lecciones")
    if request.method == "POST":
        try:
            lecciones = int(raw_lecciones) if raw_lecciones else (lecciones_programadas or 1)
        except (TypeError, ValueError):
            lecciones = lecciones_programadas or 1
    else:
        # En GET, usar el horario para poblar el campo por defecto y evitar
        # arrastre de parámetros antiguos en la URL.
        lecciones = lecciones_programadas or 1
    lecciones = max(1, lecciones)
    minuta = (request.POST.get("minuta", "") if request.method == "POST" else "").strip()

    # ── POST: guardar sesión ─────────────────────────────────────────────
    if request.method == "POST":
        periodo = _infer_periodo(asignacion, fecha)
        inst_id = asignacion.subarea_curso.institucion_id
        accion = (request.POST.get("accion") or "guardar").strip().lower()

        if accion == "eliminar":
            sesion_del = (
                AsistenciaSesion.objects
                .filter(docente_asignacion=asignacion, fecha=fecha)
                .order_by("id")
                .first()
            )
            if sesion_del:
                sesion_del.delete()
                messages.success(request, f"Asistencia del {fecha.strftime('%d/%m/%Y')} eliminada correctamente.")
            else:
                messages.warning(request, "No existe asistencia registrada para esa fecha.")
            return redirect(f"{request.path}?fecha={fecha}")

        try:
            with transaction.atomic():
                sesion, _ = AsistenciaSesion.objects.get_or_create(
                    docente_asignacion=asignacion,
                    fecha=fecha,
                    sesion_numero=1,
                    defaults={
                        "periodo": periodo,
                        "institucion_id": inst_id,
                        "curso_lectivo": asignacion.curso_lectivo,
                        "lecciones": lecciones,
                        "created_by": request.user,
                    },
                )
                sesion.periodo = periodo
                sesion.lecciones = lecciones
                sesion.minuta = minuta[:1000]
                sesion.save(update_fields=["periodo", "lecciones", "minuta", "updated_at"])
                matriculas = _get_estudiantes(asignacion)
                bulk_create = []
                bulk_update = []
                existing = {r.estudiante_id: r for r in sesion.registros.all()}

                for m in matriculas:
                    est_id = m.estudiante_id
                    raw_estado = request.POST.get(f"estado_{est_id}", AsistenciaRegistro.PRESENTE)
                    estado = raw_estado if raw_estado in dict(AsistenciaRegistro.ESTADO_CHOICES) else AsistenciaRegistro.PRESENTE
                    obs = request.POST.get(f"obs_{est_id}", "")[:255]
                    raw_cantidad = (request.POST.get(f"inj_{est_id}", "") or "").strip().replace(",", ".")
                    cantidad = None
                    if raw_cantidad != "":
                        try:
                            cantidad = Decimal(raw_cantidad)
                        except Exception:
                            messages.error(request, f"Cantidad inválida para estudiante ID {est_id}.")
                            return redirect(f"{request.path}?fecha={fecha}")
                    try:
                        lecc_inj = _resolver_lecciones_injustificadas(
                            estado=estado,
                            lecciones_dia=lecciones,
                            cantidad_ingresada=cantidad,
                            legacy_full_day_ai=True,
                        )
                    except ValidationError as exc:
                        msg = exc.messages[0] if getattr(exc, "messages", None) else str(exc)
                        messages.error(request, f"{msg} (estudiante ID {est_id}).")
                        return redirect(f"{request.path}?fecha={fecha}")

                    if est_id in existing:
                        reg = existing[est_id]
                        reg.estado = estado
                        reg.lecciones_injustificadas = lecc_inj
                        reg.observacion = obs
                        bulk_update.append(reg)
                    else:
                        bulk_create.append(
                            AsistenciaRegistro(
                                sesion=sesion,
                                estudiante_id=est_id,
                                estado=estado,
                                lecciones_injustificadas=lecc_inj,
                                observacion=obs,
                            )
                        )

                if bulk_create:
                    AsistenciaRegistro.objects.bulk_create(bulk_create)
                if bulk_update:
                    AsistenciaRegistro.objects.bulk_update(
                        bulk_update,
                        ["estado", "lecciones_injustificadas", "observacion"],
                    )

            messages.success(
                request,
                f"✔ Asistencia del {fecha.strftime('%d/%m/%Y')} guardada ({lecciones} lecciones).",
            )
        except Exception as exc:
            messages.error(request, f"Error al guardar: {exc}")

        return redirect(f"{request.path}?fecha={fecha}")

    # ── GET ──────────────────────────────────────────────────────────────
    sesion_actual = (
        AsistenciaSesion.objects
        .filter(docente_asignacion=asignacion, fecha=fecha)
        .order_by("sesion_numero")
        .first()
    )
    if sesion_actual:
        lecciones = sesion_actual.lecciones or 1
        minuta = sesion_actual.minuta or ""
        logger.info(
            "asistencia_lecciones: sesion_existente asignacion=%s fecha=%s lecciones_guardadas=%s lecciones_horario=%s",
            asignacion.id,
            fecha,
            lecciones,
            lecciones_programadas,
        )
    else:
        logger.info(
            "asistencia_lecciones: sesion_nueva asignacion=%s fecha=%s lecciones_horario=%s lecciones_form=%s",
            asignacion.id,
            fecha,
            lecciones_programadas,
            lecciones,
        )

    # Estados guardados de la fecha seleccionada
    estados_guardados = {}
    obs_guardadas = {}
    inj_guardadas = {}
    if sesion_actual:
        for reg in sesion_actual.registros.select_related("estudiante"):
            estado = reg.estado
            if estado == "T":
                estado = AsistenciaRegistro.TARDIA_MEDIA
            estados_guardados[reg.estudiante_id] = estado
            obs_guardadas[reg.estudiante_id] = reg.observacion
            inj_val = reg.lecciones_injustificadas
            if inj_val is None:
                inj_guardadas[reg.estudiante_id] = None
            elif estado in (AsistenciaRegistro.PRESENTE, AsistenciaRegistro.AUSENTE_JUSTIFICADA):
                inj_guardadas[reg.estudiante_id] = None
            elif estado == AsistenciaRegistro.TARDIA_MEDIA:
                inj_guardadas[reg.estudiante_id] = _formatear_cantidad_asistencia(
                    Decimal(str(inj_val)) * Decimal("2")
                )
            else:
                inj_guardadas[reg.estudiante_id] = _formatear_cantidad_asistencia(inj_val)

    matriculas = _get_estudiantes(asignacion)
    estudiantes = []
    for m in matriculas:
        est = m.estudiante
        estado = estados_guardados.get(est.id, AsistenciaRegistro.PRESENTE)
        estudiantes.append({
            "id": est.id,
            "nombre": str(est),
            "identificacion": est.identificacion,
            "estado": estado,
            "lecciones_injustificadas": inj_guardadas.get(est.id),
            "observacion": obs_guardadas.get(est.id, ""),
        })
    estudiantes.sort(key=lambda e: e["nombre"])

    periodo = _infer_periodo(asignacion, fecha)
    periodos_cl = (
        PeriodoCursoLectivo.objects
        .filter(
            institucion_id=asignacion.subarea_curso.institucion_id,
            curso_lectivo=asignacion.curso_lectivo,
        )
        .select_related("periodo")
        .order_by("periodo__numero")
    )

    return render(request, "libro_docente/asistencia.html", {
        "asignacion": asignacion,
        "fecha": fecha,
        "hoy": hoy,
        "lecciones": lecciones,
        "sesion_actual": sesion_actual,
        "estudiantes": estudiantes,
        "total_estudiantes": len(estudiantes),
        "periodo": periodo,
        "periodos_cl": periodos_cl,
        "minuta": minuta,
        "lecciones_programadas": lecciones_programadas,
        "PRESENTE": AsistenciaRegistro.PRESENTE,
        "TARDIA_MEDIA": AsistenciaRegistro.TARDIA_MEDIA,
        "TARDIA_COMPLETA": AsistenciaRegistro.TARDIA_COMPLETA,
        "AI": AsistenciaRegistro.AUSENTE_INJUSTIFICADA,
        "AJ": AsistenciaRegistro.AUSENTE_JUSTIFICADA,
    })


def _obtener_asignacion_con_permiso(request, asignacion_id):
    """
    Devuelve la asignación si el usuario tiene acceso.
    Superadmin: cualquier asignación. Usuario normal: solo sus asignaciones.
    """
    profesor = _get_profesor(request)
    if request.user.is_superuser:
        return get_object_or_404(
            DocenteAsignacion.objects.select_related(
                "subarea_curso__subarea", "curso_lectivo",
                "seccion__nivel", "subgrupo__seccion__nivel",
                "centro_trabajo",
            ),
            id=asignacion_id, activo=True,
        )
    if not profesor:
        return None
    return get_object_or_404(
        DocenteAsignacion.objects.select_related(
            "subarea_curso__subarea", "curso_lectivo",
            "seccion__nivel", "subgrupo__seccion__nivel",
            "centro_trabajo",
        ),
        id=asignacion_id, docente=profesor, activo=True,
    )


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def detalle_estudiante_view(request, asignacion_id, estudiante_id):
    """
    Detalle de asistencia de un estudiante en un período.
    Historial de registros + resumen acumulado (reutiliza _calcular_resumen).
    """
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")

    inst_id = asignacion.subarea_curso.institucion_id
    periodos_cl = list(
        PeriodoCursoLectivo.objects
        .filter(
            institucion_id=inst_id,
            curso_lectivo=asignacion.curso_lectivo,
        )
        .select_related("periodo")
        .order_by("periodo__numero")
    )

    # Período seleccionado
    try:
        periodo_id = int(request.GET.get("periodo", 0))
    except (ValueError, TypeError):
        periodo_id = 0
    if not periodo_id and periodos_cl:
        p = _infer_periodo(asignacion, timezone.localdate())
        periodo_id = (p.id if p else periodos_cl[0].periodo_id)

    periodo_sel = None
    for pcl in periodos_cl:
        if pcl.periodo_id == periodo_id:
            periodo_sel = pcl.periodo
            break

    # Verificar que el estudiante pertenece al grupo
    matriculas = _get_estudiantes(asignacion)
    matricula_est = matriculas.filter(estudiante_id=estudiante_id).first()
    if not matricula_est:
        messages.error(request, "El estudiante no pertenece a este grupo.")
        url = reverse("libro_docente:resumen", args=[asignacion_id])
        if periodo_id:
            url += f"?periodo={periodo_id}"
        return redirect(url)

    estudiante = matricula_est.estudiante

    # Historial de asistencia (registros del estudiante en el período)
    historial = []
    if periodo_sel:
        sesiones = _sesiones_por_periodo(asignacion, periodo_sel)
        sesion_ids = list(sesiones.values_list("id", flat=True))
        registros = list(
            AsistenciaRegistro.objects
            .filter(sesion_id__in=sesion_ids, estudiante_id=estudiante_id)
            .select_related("sesion")
            .order_by("sesion__fecha", "sesion__sesion_numero")
        )
        for reg in registros:
            estado = reg.estado
            if estado == "T":
                estado = AsistenciaRegistro.TARDIA_MEDIA
            estado_display = dict(AsistenciaRegistro.ESTADO_CHOICES).get(estado, reg.get_estado_display())
            historial.append({
                "fecha": reg.sesion.fecha,
                "lecciones": reg.sesion.lecciones or 1,
                "estado": estado,
                "estado_display": estado_display,
                "lecciones_injustificadas": reg.lecciones_injustificadas,
                "observacion": reg.observacion or "",
            })
        # Sesiones sin registro (cuentan como AI)
        sesiones_data = {s.id: (s.fecha, s.lecciones or 1) for s in sesiones}
        registradas_sesion_ids = {r.sesion_id for r in registros}
        for sid, (f, lecciones) in sesiones_data.items():
            if sid not in registradas_sesion_ids:
                historial.append({
                    "fecha": f,
                    "lecciones": lecciones,
                    "estado": AsistenciaRegistro.AUSENTE_INJUSTIFICADA,
                    "estado_display": "Ausente injustificada",
                    "lecciones_injustificadas": lecciones,
                    "observacion": "(sin registro)",
                })
        historial.sort(key=lambda x: x["fecha"])

    # Resumen acumulado (reutilizar _calcular_resumen para un solo estudiante)
    resumen_est = None
    nombre_componente = "Asistencia"
    if periodo_sel:
        resumen = _calcular_resumen(asignacion, periodo_sel, [matricula_est])
        nombre_componente = resumen.get("nombre_componente", "Asistencia")
        for r in resumen["estudiantes"]:
            if r["estudiante"].id == estudiante_id:
                resumen_est = r
                break

    return render(request, "libro_docente/detalle_estudiante.html", {
        "asignacion": asignacion,
        "estudiante": estudiante,
        "periodos_cl": periodos_cl,
        "periodo_id": periodo_id,
        "periodo_sel": periodo_sel,
        "historial": historial,
        "resumen": resumen_est,
        "nombre_componente": nombre_componente,
        "ESTADO_CHOICES": dict(AsistenciaRegistro.ESTADO_CHOICES),
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def resumen_view(request, asignacion_id):
    """Resumen de asistencia por período para una asignación."""
    profesor = _get_profesor(request)
    if not profesor:
        messages.error(request, "No tienes perfil de docente en esta institución.")
        return redirect("libro_docente:home")

    asignacion = get_object_or_404(
        DocenteAsignacion.objects.select_related(
            "subarea_curso__subarea", "curso_lectivo",
            "seccion__nivel", "subgrupo__seccion__nivel", "centro_trabajo",
        ),
        id=asignacion_id, docente=profesor, activo=True,
    )

    inst_id = asignacion.subarea_curso.institucion_id
    periodos_cl = list(
        PeriodoCursoLectivo.objects
        .filter(
            institucion_id=inst_id,
            curso_lectivo=asignacion.curso_lectivo,
        )
        .select_related("periodo")
        .order_by("periodo__numero")
    )

    # Período seleccionado
    try:
        periodo_id = int(request.GET.get("periodo", 0))
    except (ValueError, TypeError):
        periodo_id = 0

    if not periodo_id:
        p = _infer_periodo(asignacion, timezone.localdate())
        if p:
            periodo_id = p.id
        elif periodos_cl:
            periodo_id = periodos_cl[0].periodo_id

    periodo_sel = None
    for pcl in periodos_cl:
        if pcl.periodo_id == periodo_id:
            periodo_sel = pcl.periodo
            break

    resumen = {
        "total_sesiones": 0,
        "total_lecciones_periodo": 0,
        "peso_asistencia": Decimal("0"),
        "tiene_componente": False,
        "estudiantes": [],
    }
    if periodo_sel:
        matriculas = _get_estudiantes(asignacion)
        resumen = _calcular_resumen(asignacion, periodo_sel, matriculas)

    return render(request, "libro_docente/resumen.html", {
        "asignacion": asignacion,
        "periodos_cl": periodos_cl,
        "periodo_id": periodo_id,
        "periodo_sel": periodo_sel,
        "resumen": resumen,
        "institucion_activa_id": getattr(request, "institucion_activa_id", None),
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def estudiante_consulta_view(request, asignacion_id, estudiante_id):
    """
    Vista segura de ficha de estudiante para Libro Docente.
    Solo permite ver estudiantes que pertenecen a la asignación del docente.
    """
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")

    institucion = asignacion.subarea_curso.institucion
    qs_matriculas = _get_estudiantes(asignacion)
    matricula = qs_matriculas.filter(estudiante_id=estudiante_id).select_related("estudiante").first()
    if not matricula:
        messages.error(request, "El estudiante no pertenece a esta asignación.")
        return redirect("libro_docente:resumen", asignacion_id=asignacion.id)

    estudiante = matricula.estudiante
    encargados = (
        estudiante.encargadoestudiante_set
        .select_related("persona_contacto", "parentesco")
        .all()
    )

    edad_estudiante = ""
    if estudiante.fecha_nacimiento:
        today = date.today()
        years = today.year - estudiante.fecha_nacimiento.year
        months = today.month - estudiante.fecha_nacimiento.month
        if today.day < estudiante.fecha_nacimiento.day:
            months -= 1
        if months < 0:
            years -= 1
            months += 12
        if years == 0:
            edad_estudiante = f"{months} meses"
        elif months == 0:
            edad_estudiante = f"{years} años"
        else:
            edad_estudiante = f"{years} años y {months} meses"

    plantilla = PlantillaImpresionMatricula.objects.filter(institucion=institucion).first()
    context = {
        "estudiante": estudiante,
        "matricula": matricula,
        "encargados": encargados,
        "curso_lectivo": asignacion.curso_lectivo,
        "identificacion": estudiante.identificacion,
        "error": "",
        "institucion": institucion,
        "cursos_lectivos": [asignacion.curso_lectivo],
        "instituciones": [institucion],
        "es_superusuario": request.user.is_superuser,
        "plantilla": plantilla,
        "edad_estudiante": edad_estudiante,
        "mostrar_form_busqueda": False,
    }
    return render(request, "matricula/consulta_estudiante.html", context)


# ═══════════════════════════════════════════════════════════════════════════
#  EVALUACIÓN POR INDICADORES (TAREAS / COTIDIANOS)
# ═══════════════════════════════════════════════════════════════════════════


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def actividad_list_view(request, asignacion_id):
    """
    Lista actividades de evaluación de una asignación.
    Filtra por periodo y tipo_componente (opcionales por GET).
    """
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")

    inst_id = asignacion.subarea_curso.institucion_id
    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and inst_id != inst_activa:
        messages.error(request, "No puedes acceder a actividades de otra institución.")
        return redirect("libro_docente:home")

    periodos_cl = list(
        PeriodoCursoLectivo.objects
        .filter(institucion_id=inst_id, curso_lectivo=asignacion.curso_lectivo, activo=True)
        .select_related("periodo")
        .order_by("periodo__numero")
    )

    available_tipos = _tipos_habilitados_por_esquema(asignacion)
    if not available_tipos:
        available_tipos = [ActividadEvaluacion.TAREA, ActividadEvaluacion.COTIDIANO]

    periodo_id_raw = request.GET.get("periodo")
    periodo_id = int(periodo_id_raw) if periodo_id_raw and str(periodo_id_raw).isdigit() else None
    tipo = request.GET.get("tipo", "").upper()
    orden = request.GET.get("orden", "fecha")

    if tipo not in available_tipos:
        tipo = available_tipos[0]

    if tipo in available_tipos and not periodo_id and periodos_cl:
        first_pcl = periodos_cl[0]
        return redirect(
            reverse("libro_docente:actividad_list", args=[asignacion_id])
            + f"?periodo={first_pcl.periodo_id}&tipo={tipo}"
        )

    qs = ActividadEvaluacion.objects.filter(
        docente_asignacion=asignacion,
        institucion_id=inst_id,
    ).select_related("periodo").prefetch_related("indicadores")

    if periodo_id:
        qs = qs.filter(periodo_id=periodo_id)
    qs = qs.filter(tipo_componente=tipo)

    if orden == "titulo":
        qs = qs.order_by("titulo")
    elif orden == "estado":
        qs = qs.order_by("estado", "-created_at")
    elif orden == "tipo":
        qs = qs.order_by("tipo_componente", "-created_at")
    else:
        qs = qs.order_by("-created_at")

    actividades_raw = list(qs)
    actividades = []
    for a in actividades_raw:
        es_simple = a.tipo_componente in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO)
        if es_simple:
            total_max = a.puntaje_total or 0
            total_ind = 1 if total_max else 0
        else:
            indicadores_activos = [i for i in a.indicadores.all() if i.activo]
            total_max = sum(i.escala_max for i in indicadores_activos)
            total_ind = len(indicadores_activos)
        actividades.append({
            "obj": a,
            "total_indicadores": total_ind,
            "total_maximo": total_max,
        })

    return render(request, "libro_docente/actividad_list.html", {
        "asignacion": asignacion,
        "actividades": actividades,
        "periodos_cl": periodos_cl,
        "periodo_id": str(periodo_id) if periodo_id else None,
        "tipo": tipo,
        "orden": orden,
        "available_tipos": available_tipos,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def actividad_create_view(request, asignacion_id):
    """
    Crear actividad de evaluación.
    Requiere periodo y tipo por GET.
    """
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso a esta asignación.")
        return redirect("libro_docente:home")

    inst_id = asignacion.subarea_curso.institucion_id
    if getattr(request, "institucion_activa_id", None) and inst_id != request.institucion_activa_id:
        messages.error(request, "No puedes crear actividades en otra institución.")
        return redirect("libro_docente:home")

    periodo_id = request.GET.get("periodo")
    tipo = request.GET.get("tipo", "").upper()
    tipos_habilitados = _tipos_habilitados_por_esquema(asignacion)
    if not periodo_id or tipo not in TIPOS_EVALUACION:
        messages.error(request, "Debe indicar periodo y tipo válido.")
        return redirect(reverse("libro_docente:actividad_list", args=[asignacion_id]))
    if tipo not in tipos_habilitados:
        messages.error(request, "Ese componente no está habilitado en el esquema de esta asignación.")
        return redirect(reverse("libro_docente:actividad_list", args=[asignacion_id]))

    pcl = get_object_or_404(
        PeriodoCursoLectivo.objects.filter(
            institucion_id=inst_id,
            curso_lectivo=asignacion.curso_lectivo,
            activo=True,
        ).select_related("periodo"),
        periodo_id=int(periodo_id),
    )
    periodo = pcl.periodo

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data.setdefault("tipo_componente", tipo)
        post_data.setdefault("estado", ActividadEvaluacion.BORRADOR)
        post_data.setdefault("descripcion", "")
        post_data.setdefault("alcance_estudiantes", ActividadEvaluacion.ALCANCE_TODOS)
        form = ActividadEvaluacionForm(post_data)
        if form.is_valid():
            if tipo in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO):
                disponible, pct_esquema, pct_usado = porcentaje_disponible_para_tipo(
                    asignacion, periodo.id, tipo
                )
                pct_nuevo = form.cleaned_data.get("porcentaje_actividad") or Decimal("0")
                if pct_nuevo > disponible:
                    if disponible <= 0:
                        form.add_error("porcentaje_actividad", f"Ya se completó el {pct_esquema}% permitido para {tipo.title()} en este período.")
                    else:
                        form.add_error(
                            "porcentaje_actividad",
                            f"Máximo disponible: {disponible}% (esquema {pct_esquema}%, usado {pct_usado}%).",
                        )
                    return render(request, "libro_docente/actividad_form.html", {
                        "form": form,
                        "formset": None,
                        "asignacion": asignacion,
                        "periodo": periodo,
                        "periodo_id": periodo_id,
                        "tipo": tipo,
                        "tipo_display": dict(ActividadEvaluacion.TIPO_CHOICES).get(tipo, tipo.title()),
                        "actividad": None,
                        "es_simple": True,
                    })
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.docente_asignacion = asignacion
                obj.institucion_id = inst_id
                obj.curso_lectivo = asignacion.curso_lectivo
                obj.periodo = periodo
                obj.tipo_componente = tipo
                obj.created_by = request.user
                obj.save()
            if tipo in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO):
                messages.success(request, "Actividad creada. Puede calificar ahora.")
                return redirect(reverse("libro_docente:actividad_calificar", args=[obj.id]))
            messages.success(request, "Actividad creada. Agregue indicadores a continuación.")
            return redirect(reverse("libro_docente:actividad_edit", args=[obj.id]))
    else:
        form = ActividadEvaluacionForm(initial={
            "tipo_componente": tipo,
            "estado": ActividadEvaluacion.BORRADOR,
            "alcance_estudiantes": ActividadEvaluacion.ALCANCE_TODOS,
        })

    tipo_display = dict(ActividadEvaluacion.TIPO_CHOICES).get(tipo, tipo.title())
    return render(request, "libro_docente/actividad_form.html", {
        "form": form,
        "formset": None,
        "asignacion": asignacion,
        "periodo": periodo,
        "periodo_id": periodo_id,
        "tipo": tipo,
        "tipo_display": tipo_display,
        "actividad": None,
        "es_simple": tipo in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO),
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def actividad_edit_view(request, actividad_id):
    """
    Editar actividad e indicadores.
    """
    actividad = get_object_or_404(
        ActividadEvaluacion.objects.select_related(
            "docente_asignacion__subarea_curso",
            "periodo",
            "institucion",
        ),
        pk=actividad_id,
    )

    if not puede_usuario_editar_actividad(actividad, request):
        messages.error(request, "No tienes permiso para editar esta actividad.")
        return redirect("libro_docente:home")

    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and not actividad_pertenece_a_institucion(actividad, inst_activa):
        messages.error(request, "No puedes editar actividades de otra institución.")
        return redirect("libro_docente:home")

    es_simple = actividad.tipo_componente in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO)
    formset = None
    if request.method == "POST":
        post_data = request.POST.copy()
        fecha_asignacion_raw = (request.POST.get("fecha_asignacion") or "").strip()
        fecha_entrega_raw = (request.POST.get("fecha_entrega") or "").strip()
        post_data.setdefault("tipo_componente", actividad.tipo_componente)
        if es_simple:
            post_data.setdefault("estado", actividad.estado or ActividadEvaluacion.BORRADOR)
            post_data.setdefault("descripcion", actividad.descripcion or "")
            post_data.setdefault("alcance_estudiantes", actividad.alcance_estudiantes or ActividadEvaluacion.ALCANCE_TODOS)
        form = ActividadEvaluacionForm(post_data, instance=actividad)
        if not es_simple:
            formset = IndicadorActividadFormSet(request.POST, instance=actividad)
        valid = form.is_valid() and (es_simple or (formset and formset.is_valid()))
        if valid:
            if es_simple:
                disponible, pct_esquema, pct_usado = porcentaje_disponible_para_tipo(
                    actividad.docente_asignacion,
                    actividad.periodo_id,
                    actividad.tipo_componente,
                    actividad_excluir_id=actividad.id,
                )
                pct_nuevo = form.cleaned_data.get("porcentaje_actividad") or Decimal("0")
                if pct_nuevo > disponible:
                    if disponible <= 0:
                        form.add_error("porcentaje_actividad", f"Ya se completó el {pct_esquema}% permitido en este período.")
                    else:
                        form.add_error(
                            "porcentaje_actividad",
                            f"Máximo disponible: {disponible}% (esquema {pct_esquema}%, usado {pct_usado}%).",
                        )
                else:
                    with transaction.atomic():
                        obj = form.save(commit=False)
                        if not fecha_asignacion_raw and actividad.fecha_asignacion:
                            obj.fecha_asignacion = actividad.fecha_asignacion
                        if not fecha_entrega_raw and actividad.fecha_entrega:
                            obj.fecha_entrega = actividad.fecha_entrega
                        obj.save()
                    messages.success(request, "Actividad actualizada.")
                    return redirect(reverse("libro_docente:actividad_edit", args=[actividad_id]))
            else:
                with transaction.atomic():
                    obj = form.save(commit=False)
                    if not fecha_asignacion_raw and actividad.fecha_asignacion:
                        obj.fecha_asignacion = actividad.fecha_asignacion
                    if not fecha_entrega_raw and actividad.fecha_entrega:
                        obj.fecha_entrega = actividad.fecha_entrega
                    obj.save()
                    formset.instance = obj
                    formset.save()
                    crear_copia_adecuacion = (
                        request.POST.get("crear_copia_adecuacion") == "1"
                        and actividad.tipo_componente in (ActividadEvaluacion.TAREA, ActividadEvaluacion.COTIDIANO)
                        and actividad.alcance_estudiantes != ActividadEvaluacion.ALCANCE_ADECUACION
                    )
                    if crear_copia_adecuacion:
                        ids_adecuacion = _get_ids_adecuacion(actividad.docente_asignacion)
                        if ids_adecuacion:
                            copia = duplicar_actividad(
                                actividad,
                                titulo_nuevo=f"{actividad.titulo} (Adecuación)",
                            )
                            copia.alcance_estudiantes = ActividadEvaluacion.ALCANCE_ADECUACION
                            copia.created_by = request.user
                            copia.save(update_fields=["alcance_estudiantes", "created_by"])
                messages.success(request, "Actividad actualizada.")
                return redirect(reverse("libro_docente:actividad_edit", args=[actividad_id]))
    else:
        form = ActividadEvaluacionForm(instance=actividad)
        if not es_simple:
            formset = IndicadorActividadFormSet(instance=actividad)

    total_maximo = (actividad.puntaje_total or 0) if es_simple else calcular_total_maximo_actividad(actividad)

    return render(request, "libro_docente/actividad_form.html", {
        "form": form,
        "formset": formset,
        "asignacion": actividad.docente_asignacion,
        "periodo": actividad.periodo,
        "periodo_id": actividad.periodo_id,
        "tipo": actividad.tipo_componente,
        "tipo_display": actividad.get_tipo_componente_display(),
        "actividad": actividad,
        "total_maximo": total_maximo,
        "es_simple": es_simple,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def actividad_delete_view(request, actividad_id):
    """
    Eliminar actividad (con confirmación por POST).
    """
    actividad = get_object_or_404(
        ActividadEvaluacion.objects.select_related("docente_asignacion"),
        pk=actividad_id,
    )

    if not puede_usuario_editar_actividad(actividad, request):
        messages.error(request, "No tienes permiso para eliminar esta actividad.")
        return redirect("libro_docente:home")

    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and not actividad_pertenece_a_institucion(actividad, inst_activa):
        messages.error(request, "No puedes eliminar actividades de otra institución.")
        return redirect("libro_docente:home")

    asignacion_id = actividad.docente_asignacion_id

    if request.method == "POST":
        actividad.delete()
        messages.success(request, "Actividad eliminada.")
        return redirect(reverse("libro_docente:actividad_list", args=[asignacion_id]))

    return render(request, "libro_docente/actividad_confirm_delete.html", {
        "actividad": actividad,
        "asignacion": actividad.docente_asignacion,
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def actividad_duplicar_view(request, actividad_id):
    """
    Duplicar actividad con sus indicadores (sin puntajes).
    """
    actividad = get_object_or_404(
        ActividadEvaluacion.objects.select_related("docente_asignacion").prefetch_related("indicadores"),
        pk=actividad_id,
    )

    if not puede_usuario_editar_actividad(actividad, request):
        messages.error(request, "No tienes permiso para duplicar esta actividad.")
        return redirect("libro_docente:home")

    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and not actividad_pertenece_a_institucion(actividad, inst_activa):
        messages.error(request, "No puedes duplicar actividades de otra institución.")
        return redirect("libro_docente:home")

    if actividad.tipo_componente in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO):
        disponible, pct_esquema, pct_usado = porcentaje_disponible_para_tipo(
            actividad.docente_asignacion,
            actividad.periodo_id,
            actividad.tipo_componente,
        )
        pct_actividad = actividad.porcentaje_actividad or Decimal("0")
        if pct_actividad > disponible:
            messages.error(
                request,
                f"No se puede duplicar: disponible {disponible}% (esquema {pct_esquema}%, usado {pct_usado}%).",
            )
            return redirect(reverse("libro_docente:actividad_list", args=[actividad.docente_asignacion_id]))

    nueva = duplicar_actividad(actividad)
    nueva.created_by = request.user
    nueva.save(update_fields=["created_by"])
    messages.success(request, f"Copia creada: «{nueva.titulo}».")
    return redirect(reverse("libro_docente:actividad_edit", args=[nueva.id]))


def _asignaciones_destino_para_copiar(actividad, request):
    """
    Otras asignaciones del mismo docente, misma materia (subarea), mismo curso_lectivo,
    misma institución, distintas del grupo actual. Solo las que tienen el periodo.
    """
    asignacion_actual = actividad.docente_asignacion
    profesor = _get_profesor(request)
    if not profesor:
        return []
    if request.user.is_superuser:
        qs = DocenteAsignacion.objects.filter(activo=True)
    else:
        qs = DocenteAsignacion.objects.filter(docente=profesor, activo=True)
    inst_id = asignacion_actual.subarea_curso.institucion_id
    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and inst_id != inst_activa:
        return []
    subarea_id = asignacion_actual.subarea_curso.subarea_id
    curso_lectivo_id = asignacion_actual.curso_lectivo_id
    periodo = actividad.periodo
    pcl_existe = PeriodoCursoLectivo.objects.filter(
        institucion_id=inst_id,
        curso_lectivo_id=curso_lectivo_id,
        periodo=periodo,
        activo=True,
    ).exists()
    if not pcl_existe:
        return []
    return list(
        qs.filter(
            subarea_curso__institucion_id=inst_id,
            subarea_curso__subarea_id=subarea_id,
            curso_lectivo_id=curso_lectivo_id,
        )
        .exclude(id=asignacion_actual.id)
        .select_related("subarea_curso__subarea", "seccion", "subgrupo__seccion")
        .order_by("seccion__nivel__numero", "seccion__numero", "subgrupo__letra")
    )


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def actividad_copiar_a_grupos_view(request, actividad_id):
    """
    Copiar actividad a otros grupos/asignaciones (misma materia, mismo curso lectivo).
    """
    actividad = get_object_or_404(
        ActividadEvaluacion.objects.select_related(
            "docente_asignacion__subarea_curso",
            "periodo",
        ).prefetch_related("indicadores"),
        pk=actividad_id,
    )

    if not puede_usuario_editar_actividad(actividad, request):
        messages.error(request, "No tienes permiso para copiar esta actividad.")
        return redirect("libro_docente:home")

    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and not actividad_pertenece_a_institucion(actividad, inst_activa):
        messages.error(request, "No puedes copiar actividades de otra institución.")
        return redirect("libro_docente:home")

    asignaciones_destino = _asignaciones_destino_para_copiar(actividad, request)

    if request.method == "POST":
        ids_str = request.POST.getlist("asignacion_id")
        asignacion_ids = [int(x) for x in ids_str if str(x).isdigit()]
        if asignacion_ids:
            creadas = copiar_actividad_a_asignaciones(actividad, asignacion_ids, created_by=request.user)
            if creadas:
                messages.success(
                    request,
                    f"Actividad copiada a {len(creadas)} grupo(s): «{actividad.titulo}».",
                )
            else:
                messages.warning(request, "No se pudo copiar a ningún grupo. Verifique los destinos.")
        else:
            messages.warning(request, "Seleccione al menos un grupo destino.")
        return redirect(reverse("libro_docente:actividad_list", args=[actividad.docente_asignacion_id]))

    return render(request, "libro_docente/actividad_copiar_a_grupos.html", {
        "actividad": actividad,
        "asignacion": actividad.docente_asignacion,
        "asignaciones_destino": asignaciones_destino,
        "es_simple": actividad.tipo_componente in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO),
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def actividad_calificar_view(request, actividad_id):
    """
    Grilla de calificación por estudiante e indicador.
    GET: muestra grilla con puntajes existentes.
    POST: guardado masivo de puntajes.
    """
    from decimal import Decimal

    actividad = get_object_or_404(
        ActividadEvaluacion.objects.select_related(
            "docente_asignacion__subarea_curso",
            "periodo",
        ).prefetch_related("indicadores"),
        pk=actividad_id,
    )

    if not puede_usuario_editar_actividad(actividad, request):
        messages.error(request, "No tienes permiso para calificar esta actividad.")
        return redirect("libro_docente:home")

    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and not actividad_pertenece_a_institucion(actividad, inst_activa):
        messages.error(request, "No puedes calificar actividades de otra institución.")
        return redirect("libro_docente:home")

    asignacion = actividad.docente_asignacion
    matriculas = _get_estudiantes_para_actividad(asignacion, actividad)
    es_simple = actividad.tipo_componente in (ActividadEvaluacion.PRUEBA, ActividadEvaluacion.PROYECTO)
    if es_simple:
        return _calificar_simple(request, actividad, asignacion, matriculas)
    indicadores = list(actividad.indicadores.filter(activo=True).order_by("orden", "id"))
    total_maximo = calcular_total_maximo_actividad(actividad)

    # Cargar puntajes existentes
    puntajes_existentes = {}
    observaciones_existentes = {}
    if indicadores and matriculas:
        ind_ids = [ind.id for ind in indicadores]
        est_ids = [m.estudiante_id for m in matriculas]
        for p in PuntajeIndicador.objects.filter(
            indicador_id__in=ind_ids,
            estudiante_id__in=est_ids,
        ).select_related("indicador"):
            if p.puntaje_obtenido is not None:
                puntajes_existentes[(p.indicador_id, p.estudiante_id)] = p.puntaje_obtenido
        for o in ObservacionActividadEstudiante.objects.filter(
            actividad=actividad,
            estudiante_id__in=est_ids,
        ):
            observaciones_existentes[o.estudiante_id] = o.observacion or ""

    # POST: guardar puntajes
    if request.method == "POST":
        datos = {}
        observaciones_post = {}
        for ind in indicadores:
            for m in matriculas:
                key = f"p_{ind.id}_{m.estudiante_id}"
                val = request.POST.get(key)
                if val is not None:
                    datos[(ind.id, m.estudiante_id)] = val
        for m in matriculas:
            obs_key = f"obs_{m.estudiante_id}"
            observaciones_post[m.estudiante_id] = (request.POST.get(obs_key, "") or "").strip()
        with transaction.atomic():
            guardados, errores = guardar_puntajes_masivo(
                actividad,
                [m.estudiante_id for m in matriculas],
                datos,
                indicadores_ids={ind.id for ind in indicadores},
            )
            for est_id, obs_text in observaciones_post.items():
                if obs_text:
                    ObservacionActividadEstudiante.objects.update_or_create(
                        actividad=actividad,
                        estudiante_id=est_id,
                        defaults={"observacion": obs_text},
                    )
                else:
                    ObservacionActividadEstudiante.objects.filter(
                        actividad=actividad,
                        estudiante_id=est_id,
                    ).delete()
        if errores:
            for e in errores:
                messages.error(request, e)
        elif guardados > 0:
            messages.success(request, f"Se guardaron {guardados} puntaje(s).")
        else:
            messages.info(request, "No hubo cambios que guardar.")
        return redirect(reverse("libro_docente:actividad_calificar", args=[actividad_id]))

    # Construir filas para la grilla: cada fila tiene indicador_puntajes = [(ind, valor), ...]
    filas = []
    for m in matriculas:
        est = m.estudiante
        indicador_puntajes = [
            (ind, puntajes_existentes.get((ind.id, est.id)))
            for ind in indicadores
        ]
        total_obt = sum(
            (p or Decimal("0")) for _, p in indicador_puntajes
        )
        pct = (
            (total_obt / total_maximo * Decimal("100"))
            if total_maximo and total_maximo > 0
            else Decimal("0")
        )
        filas.append({
            "matricula": m,
            "estudiante": est,
            "indicador_puntajes": indicador_puntajes,
            "total_obtenido": total_obt,
            "porcentaje_logro": pct,
            "observacion_tarea": observaciones_existentes.get(est.id, ""),
        })

    # Orden por apellido (primer_apellido, segundo_apellido, nombres)
    filas.sort(
        key=lambda f: (
            (f["estudiante"].primer_apellido or "").upper(),
            (f["estudiante"].segundo_apellido or "").upper(),
            (f["estudiante"].nombres or "").upper(),
        )
    )

    return render(request, "libro_docente/calificacion.html", {
        "actividad": actividad,
        "asignacion": asignacion,
        "indicadores": indicadores,
        "filas": filas,
        "total_maximo": total_maximo,
        "total_estudiantes": len(filas),
    })


def _calificar_simple(request, actividad, asignacion, matriculas):
    puntaje_total = actividad.puntaje_total or Decimal("0")
    porcentaje_actividad = actividad.porcentaje_actividad or Decimal("0")
    if puntaje_total <= 0:
        messages.error(request, "Defina un valor en puntos válido antes de calificar.")
        return redirect(reverse("libro_docente:actividad_edit", args=[actividad.id]))

    est_ids = [m.estudiante_id for m in matriculas]
    puntajes_existentes = {
        p.estudiante_id: p.puntos_obtenidos
        for p in PuntajeSimple.objects.filter(actividad=actividad, estudiante_id__in=est_ids)
    }

    if request.method == "POST":
        errores = []
        cambios = []
        with transaction.atomic():
            for m in matriculas:
                key = f"ps_{m.estudiante_id}"
                raw = (request.POST.get(key, "") or "").strip()
                if raw == "":
                    cambios.append(("delete", m.estudiante_id, None))
                    continue
                try:
                    puntos = Decimal(raw.replace(",", "."))
                except Exception:
                    errores.append(f"{m.estudiante}: valor inválido.")
                    continue
                if puntos < 0 or puntos > puntaje_total or puntos != puntos.to_integral_value():
                    errores.append(
                        f"{m.estudiante}: puntaje fuera de rango (0 a {int(puntaje_total)})."
                    )
                    continue
                cambios.append(("upsert", m.estudiante_id, puntos))
            if errores:
                transaction.set_rollback(True)
            else:
                for op, est_id, pts in cambios:
                    if op == "delete":
                        PuntajeSimple.objects.filter(actividad=actividad, estudiante_id=est_id).delete()
                    else:
                        PuntajeSimple.objects.update_or_create(
                            actividad=actividad,
                            estudiante_id=est_id,
                            defaults={"puntos_obtenidos": pts},
                        )
        if errores:
            for e in errores[:5]:
                messages.error(request, e)
            if len(errores) > 5:
                messages.error(request, f"Se detectaron {len(errores)} errores de validación.")
            return redirect(reverse("libro_docente:actividad_calificar", args=[actividad.id]))
        messages.success(request, "Puntajes guardados.")
        return redirect(reverse("libro_docente:actividad_calificar", args=[actividad.id]))

    filas = []
    for m in matriculas:
        est = m.estudiante
        puntos = puntajes_existentes.get(est.id)
        nota = (puntos / puntaje_total * Decimal("100")) if (puntos is not None and puntaje_total > 0) else None
        porcentaje = (puntos / puntaje_total * porcentaje_actividad) if (puntos is not None and puntaje_total > 0) else None
        filas.append({
            "matricula": m,
            "estudiante": est,
            "puntos": puntos,
            "nota": nota,
            "porcentaje": porcentaje,
        })

    filas.sort(
        key=lambda f: (
            (f["estudiante"].primer_apellido or "").upper(),
            (f["estudiante"].segundo_apellido or "").upper(),
            (f["estudiante"].nombres or "").upper(),
        )
    )
    return render(request, "libro_docente/calificacion_simple.html", {
        "actividad": actividad,
        "asignacion": asignacion,
        "filas": filas,
        "puntaje_total": puntaje_total,
        "porcentaje_actividad": porcentaje_actividad,
        "total_estudiantes": len(filas),
    })


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def resumen_evaluacion_view(request, asignacion_id):
    """
    Resumen acumulado por componente (TAREAS / COTIDIANOS) por estudiante.
    Muestra obtenidos, máximos, % logro, % componente y aporte a nota final.
    """
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso.")
        return redirect("libro_docente:home")

    inst_id = asignacion.subarea_curso.institucion_id
    inst_activa = getattr(request, "institucion_activa_id", None)
    if inst_activa and inst_id != inst_activa:
        messages.error(request, "No puedes acceder a otra institución.")
        return redirect("libro_docente:home")

    periodos_cl = list(
        PeriodoCursoLectivo.objects
        .filter(institucion_id=inst_id, curso_lectivo=asignacion.curso_lectivo, activo=True)
        .select_related("periodo")
        .order_by("periodo__numero")
    )

    periodo_id_raw = request.GET.get("periodo")
    periodo_id = int(periodo_id_raw) if periodo_id_raw and str(periodo_id_raw).isdigit() else None
    tipo = request.GET.get("tipo", "").upper()
    if tipo not in TIPOS_EVALUACION:
        tipo = None
    if not periodo_id and periodos_cl:
        periodo_id = periodos_cl[0].periodo_id

    matriculas = _get_estudiantes(asignacion)
    filas = []
    filas_general = []
    has_proyecto = ActividadEvaluacion.PROYECTO in _tipos_habilitados_por_esquema(asignacion)
    if periodo_id:
        filas = calcular_resumen_evaluacion_completo(asignacion, periodo_id, matriculas)
        filas_general = _construir_resumen_general(asignacion, periodo_id, matriculas, filas)
        periodo_sel = (
            PeriodoCursoLectivo.objects
            .filter(
                institucion_id=asignacion.subarea_curso.institucion_id,
                curso_lectivo=asignacion.curso_lectivo,
                periodo_id=periodo_id,
                activo=True,
            )
            .select_related("periodo")
            .first()
        )
        asistencia_map = {}
        if periodo_sel:
            resumen_asis = _calcular_resumen(asignacion, periodo_sel.periodo, matriculas)
            asistencia_map = {r["estudiante"].id: r["aporte_real"] for r in resumen_asis.get("estudiantes", [])}
        for f in filas:
            aporte_asistencia = asistencia_map.get(f["estudiante"].id, Decimal("0"))
            f["asistencia"] = {"aporte": aporte_asistencia}
            f["nota_final"] = (
                (f.get("tareas", {}) or {}).get("aporte", Decimal("0"))
                + (f.get("cotidianos", {}) or {}).get("aporte", Decimal("0"))
                + (f.get("pruebas", {}) or {}).get("aporte", Decimal("0"))
                + (f.get("proyectos", {}) or {}).get("aporte", Decimal("0"))
                + aporte_asistencia
            )

    return render(request, "libro_docente/resumen_evaluacion.html", {
        "asignacion": asignacion,
        "periodos_cl": periodos_cl,
        "periodo_id": str(periodo_id) if periodo_id else None,
        "tipo": tipo,
        "filas": filas,
        "filas_general": filas_general,
        "has_proyecto": has_proyecto,
    })


def _construir_resumen_general(asignacion, periodo_id, matriculas, filas_eval):
    """
    Resumen final por estudiante para exportación.
    """
    periodo_sel = (
        PeriodoCursoLectivo.objects
        .filter(
            institucion_id=asignacion.subarea_curso.institucion_id,
            curso_lectivo=asignacion.curso_lectivo,
            periodo_id=periodo_id,
            activo=True,
        )
        .select_related("periodo")
        .first()
    )
    asist_por_est = {}
    if periodo_sel:
        resumen_asis = _calcular_resumen(asignacion, periodo_sel.periodo, matriculas)
        asist_por_est = {r["estudiante"].id: r["aporte_real"] for r in resumen_asis.get("estudiantes", [])}

    filas_por_est = {f["estudiante"].id: f for f in filas_eval}
    rows = []
    for m in matriculas:
        est = m.estudiante
        fe = filas_por_est.get(est.id, {})
        rows.append({
            "id": est.identificacion,
            "nombre": f"{est.primer_apellido} {est.segundo_apellido or ''} {est.nombres}".strip(),
            "cotidiano": (fe.get("cotidianos", {}) or {}).get("aporte", Decimal("0")),
            "tareas": (fe.get("tareas", {}) or {}).get("aporte", Decimal("0")),
            "pruebas": (fe.get("pruebas", {}) or {}).get("aporte", Decimal("0")),
            "proyecto": (fe.get("proyectos", {}) or {}).get("aporte", Decimal("0")),
            "asistencia": asist_por_est.get(est.id, Decimal("0")),
        })
    return rows


def _to_2_dec(valor):
    if valor is None:
        return Decimal("0.00")
    if not isinstance(valor, Decimal):
        valor = Decimal(str(valor))
    return valor.quantize(Decimal("0.01"))


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def resumen_general_export_xlsx(request, asignacion_id):
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso.")
        return redirect("libro_docente:home")
    if openpyxl is None:
        messages.error(request, "openpyxl no está disponible en el servidor.")
        return redirect(reverse("libro_docente:resumen_evaluacion", args=[asignacion_id]))

    periodo_id_raw = request.GET.get("periodo")
    periodo_id = int(periodo_id_raw) if periodo_id_raw and str(periodo_id_raw).isdigit() else None
    if not periodo_id:
        messages.error(request, "Debe seleccionar período para exportar.")
        return redirect(reverse("libro_docente:resumen_evaluacion", args=[asignacion_id]))

    matriculas = _get_estudiantes(asignacion)
    filas = calcular_resumen_evaluacion_completo(asignacion, periodo_id, matriculas)
    filas_general = _construir_resumen_general(asignacion, periodo_id, matriculas, filas)
    has_proyecto = ActividadEvaluacion.PROYECTO in _tipos_habilitados_por_esquema(asignacion)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen General"
    headers = ["id", "Nombre", "Trabajo cotidiano", "Tareas", "Pruebas"]
    if has_proyecto:
        headers.append("Proyecto")
    headers.append("Asistencia")
    ws.append(headers)
    for r in filas_general:
        row = [
            r["id"],
            r["nombre"],
            float(_to_2_dec(r["cotidiano"])),
            float(_to_2_dec(r["tareas"])),
            float(_to_2_dec(r["pruebas"])),
        ]
        if has_proyecto:
            row.append(float(_to_2_dec(r["proyecto"])))
        row.append(float(_to_2_dec(r["asistencia"])))
        ws.append(row)

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="resumen_general_{asignacion_id}_{periodo_id}.xlsx"'
    wb.save(resp)
    return resp


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def resumen_general_export_csv(request, asignacion_id):
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso.")
        return redirect("libro_docente:home")

    periodo_id_raw = request.GET.get("periodo")
    periodo_id = int(periodo_id_raw) if periodo_id_raw and str(periodo_id_raw).isdigit() else None
    if not periodo_id:
        messages.error(request, "Debe seleccionar período para exportar.")
        return redirect(reverse("libro_docente:resumen_evaluacion", args=[asignacion_id]))

    matriculas = _get_estudiantes(asignacion)
    filas = calcular_resumen_evaluacion_completo(asignacion, periodo_id, matriculas)
    filas_general = _construir_resumen_general(asignacion, periodo_id, matriculas, filas)
    has_proyecto = ActividadEvaluacion.PROYECTO in _tipos_habilitados_por_esquema(asignacion)

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="resumen_general_{asignacion_id}_{periodo_id}.csv"'
    writer = csv.writer(resp)
    headers = ["id", "Nombre", "Trabajo cotidiano", "Tareas", "Pruebas"]
    if has_proyecto:
        headers.append("Proyecto")
    headers.append("Asistencia")
    writer.writerow(headers)
    for r in filas_general:
        row = [
            r["id"],
            r["nombre"],
            f"{_to_2_dec(r['cotidiano']):.2f}",
            f"{_to_2_dec(r['tareas']):.2f}",
            f"{_to_2_dec(r['pruebas']):.2f}",
        ]
        if has_proyecto:
            row.append(f"{_to_2_dec(r['proyecto']):.2f}")
        row.append(f"{_to_2_dec(r['asistencia']):.2f}")
        writer.writerow(row)
    return resp


@login_required
@permission_required("libro_docente.access_libro_docente", raise_exception=True)
def resumen_estudiante_detalle_view(request, asignacion_id, estudiante_id):
    """
    Detalle del resumen por estudiante: desglose por actividad.
    """
    asignacion = _obtener_asignacion_con_permiso(request, asignacion_id)
    if asignacion is None:
        messages.error(request, "No tienes acceso.")
        return redirect("libro_docente:home")

    matriculas = _get_estudiantes(asignacion)
    matricula = next((m for m in matriculas if m.estudiante_id == estudiante_id), None)
    if not matricula:
        messages.error(request, "El estudiante no pertenece a este grupo.")
        return redirect(reverse("libro_docente:resumen_evaluacion", args=[asignacion_id]))

    periodo_id_raw = request.GET.get("periodo")
    periodo_id = int(periodo_id_raw) if periodo_id_raw and str(periodo_id_raw).isdigit() else None
    tipo = request.GET.get("tipo", "").upper()
    if tipo not in TIPOS_EVALUACION:
        tipo = None
    inst_id = asignacion.subarea_curso.institucion_id
    periodos_cl = list(
        PeriodoCursoLectivo.objects
        .filter(institucion_id=inst_id, curso_lectivo=asignacion.curso_lectivo, activo=True)
        .select_related("periodo")
        .order_by("periodo__numero")
    )
    if not periodo_id and periodos_cl:
        periodo_id = periodos_cl[0].periodo_id

    from decimal import Decimal
    from .services import calcular_resumen_componente_estudiante

    _empty_resumen = {
        "puntos_obtenidos": Decimal("0"),
        "puntos_maximos": Decimal("0"),
        "porcentaje_logro": Decimal("0"),
        "porcentaje_componente": Decimal("0"),
        "aporte": Decimal("0"),
        "detalle_actividades": [],
    }
    tareas = (
        calcular_resumen_componente_estudiante(
            asignacion, periodo_id, ActividadEvaluacion.TAREA, estudiante_id
        )
        if periodo_id
        else _empty_resumen
    )
    cotidianos = (
        calcular_resumen_componente_estudiante(
            asignacion, periodo_id, ActividadEvaluacion.COTIDIANO, estudiante_id
        )
        if periodo_id
        else _empty_resumen
    )
    pruebas = (
        calcular_resumen_componente_estudiante(
            asignacion, periodo_id, ActividadEvaluacion.PRUEBA, estudiante_id
        )
        if periodo_id
        else _empty_resumen
    )
    proyectos = (
        calcular_resumen_componente_estudiante(
            asignacion, periodo_id, ActividadEvaluacion.PROYECTO, estudiante_id
        )
        if periodo_id
        else _empty_resumen
    )

    return render(request, "libro_docente/resumen_estudiante_detalle.html", {
        "asignacion": asignacion,
        "estudiante": matricula.estudiante,
        "periodos_cl": periodos_cl,
        "periodo_id": str(periodo_id) if periodo_id else None,
        "tipo": tipo,
        "tareas": tareas,
        "cotidianos": cotidianos,
        "pruebas": pruebas,
        "proyectos": proyectos,
    })
